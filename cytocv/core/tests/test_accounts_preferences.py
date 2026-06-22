"""Tests for account preference and account-area safeguards."""

from __future__ import annotations

import csv
import json
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from tempfile import TemporaryDirectory
from types import SimpleNamespace
from unittest.mock import patch
from uuid import uuid4

from django.contrib.auth import get_user_model
from django.contrib.auth.models import AnonymousUser
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from accounts.preferences import (
    get_user_preferences,
    normalize_preferences_payload,
    should_auto_save_experiments,
    update_user_preferences,
)
from core.channel_roles import (
    CHANNEL_ROLE_BLUE,
    CHANNEL_ROLE_DIC,
    CHANNEL_ROLE_GREEN,
    CHANNEL_ROLE_RED,
)
from core.models import (
    CellStatistics,
    DVLayerTifPreview,
    SegmentedImage,
    UploadedImage,
    get_guest_user,
)
from core.scale import apply_manual_override_scale, build_scale_info
from core.services.analysis_context import normalize_analysis_config_snapshot
from core.services.signal_quantification import (
    resolve_effective_alternate_nucleus_detection,
    resolve_signal_quantification_selection,
)
from core.services.stat_export_selection import USER_SELECTABLE_TABLE_FIELDS
from core.stats_plugins import PLUGIN_DEFINITIONS


CORE_STATIC_ROOT = Path(__file__).resolve().parents[1] / "static"


def _frontend_static_text(relative_path: str) -> str:
    return (CORE_STATIC_ROOT / relative_path).read_text(encoding="utf-8")


class PreferenceNormalizationTests(TestCase):
    def test_new_account_preferences_default_signal_modes_to_balanced(self):
        user = get_user_model().objects.create_user(
            email="new-balanced-defaults@example.com",
            password="TestPass123!",
        )

        defaults = get_user_preferences(user)["experiment_defaults"]

        self.assertEqual(defaults["nuclear_cell_pair_contour_mode"], "balanced")
        self.assertEqual(defaults["green_dot_split_mode"], "balanced")
        self.assertEqual(defaults["red_dot_split_mode"], "balanced")

    def test_existing_account_aggressive_signal_modes_are_preserved(self):
        user = get_user_model().objects.create_user(
            email="existing-aggressive-defaults@example.com",
            password="TestPass123!",
        )
        user.config = {
            "preferences": {
                "experiment_defaults": {
                    "nuclear_cell_pair_contour_mode": "aggressive",
                    "green_dot_split_mode": "aggressive",
                    "red_dot_split_mode": "aggressive",
                }
            }
        }
        user.save(update_fields=["config"])

        defaults = get_user_preferences(user)["experiment_defaults"]

        self.assertEqual(defaults["nuclear_cell_pair_contour_mode"], "aggressive")
        self.assertEqual(defaults["green_dot_split_mode"], "aggressive")
        self.assertEqual(defaults["red_dot_split_mode"], "aggressive")

    def test_existing_account_legacy_scaled_nuclear_mode_is_preserved(self):
        normalized = normalize_preferences_payload(
            {
                "experiment_defaults": {
                    "use_legacy_nuclear_cell_pair_pipeline": "true",
                }
            }
        )

        self.assertTrue(
            normalized["experiment_defaults"]["use_legacy_nuclear_cell_pair_pipeline"]
        )

    def test_default_payload_uses_expected_plugin_defaults(self):
        normalized = normalize_preferences_payload({})
        defaults = normalized["experiment_defaults"]
        self.assertEqual(
            defaults["selected_plugins"],
            [
                "PunctaDistance",
                "CENDot",
                "Biorientation",
                "GreenRedIntensity",
            ],
        )
        self.assertTrue(defaults["signal_quantification_enabled"])
        self.assertEqual(defaults["signal_quantification_mode"], "puncta_distance")
        self.assertTrue(defaults["puncta_contour_intensity_enabled"])
        self.assertTrue(defaults["alternate_nucleus_detection_enabled"])
        self.assertEqual(defaults["puncta_line_mode"], "red_puncta")
        self.assertEqual(defaults["nuclear_cell_pair_mode"], "green_nucleus")
        self.assertEqual(defaults["nuclear_cell_pair_contour_mode"], "balanced")
        self.assertFalse(defaults["use_legacy_nuclear_cell_pair_pipeline"])
        self.assertTrue(defaults["green_dot_split_enabled"])
        self.assertEqual(defaults["green_dot_split_mode"], "balanced")
        self.assertTrue(defaults["red_dot_split_enabled"])
        self.assertEqual(defaults["red_dot_split_mode"], "balanced")
        self.assertNotIn("puncta_source_contour_count_filter", defaults)
        self.assertTrue(defaults["use_metadata_scale"])
        self.assertEqual(defaults["spatial_stats_unit"], "px")
        self.assertTrue(defaults["use_metadata_channel_order"])
        self.assertEqual(
            defaults["fallback_channel_order"],
            [CHANNEL_ROLE_DIC, CHANNEL_ROLE_BLUE, CHANNEL_ROLE_GREEN, CHANNEL_ROLE_RED],
        )
        self.assertTrue(normalized["show_saved_file_channels"])
        self.assertTrue(normalized["show_saved_file_scales"])
        self.assertTrue(normalized["sidebar_starts_open"])
        self.assertTrue(normalized["confirm_cell_deletion"])
        self.assertTrue(normalized["confirm_multi_cell_deletion"])
        self.assertEqual(normalized["sidebar_spatial_stats_unit"], "px")
        self.assertEqual(normalized["main_image_channel"], "")
        self.assertEqual(
            normalized["default_puncta_source_contour_count_filter"],
            "all",
        )

    def test_normalize_preferences_filters_invalid_values(self):
        normalized = normalize_preferences_payload(
            {
                "experiment_defaults": {
                    "selected_plugins": ["PunctaDistance", "Unknown"],
                    "module_enabled": "true",
                    "enforce_layer_count": "true",
                    "enforce_wavelengths": "false",
                    "manual_required_channels": ["DIC", "BAD"],
                    "puncta_line_width": "-5",
                    "cen_dot_distance": "abc",
                    "cen_dot_collinearity_threshold": "-1",
                    "puncta_line_mode": "bad_mode",
                    "nuclear_cell_pair_mode": "bad_mode",
                    "nuclear_cell_pair_contour_mode": "bad_mode",
                    "green_dot_split_mode": "bad_mode",
                    "red_dot_split_enabled": "off",
                    "red_dot_split_mode": "bad_mode",
                    "puncta_source_contour_count_filter": "bad_filter",
                    "puncta_line_width_unit": "um",
                    "cen_dot_distance_unit": "px",
                    "microns_per_pixel": "0",
                    "use_metadata_scale": "off",
                    "use_metadata_channel_order": "off",
                    "fallback_channel_order": [
                        "Green",
                        "DIC",
                        "Red",
                        "Blue",
                    ],
                    "spatial_stats_unit": "bad_unit",
                },
                "auto_save_experiments": "off",
                "show_saved_file_scales": "off",
                "confirm_cell_deletion": "off",
                "confirm_multi_cell_deletion": "off",
                "main_image_channel": "invalid",
                "default_puncta_source_contour_count_filter": "bad_filter",
            }
        )

        defaults = normalized["experiment_defaults"]
        self.assertEqual(defaults["selected_plugins"], ["PunctaDistance"])
        self.assertTrue(defaults["signal_quantification_enabled"])
        self.assertEqual(defaults["signal_quantification_mode"], "puncta_distance")
        self.assertFalse(defaults["puncta_contour_intensity_enabled"])
        self.assertTrue(defaults["module_enabled"])
        self.assertTrue(defaults["enforce_layer_count"])
        self.assertFalse(defaults["enforce_wavelengths"])
        self.assertEqual(defaults["manual_required_channels"], ["DIC"])
        self.assertEqual(defaults["puncta_line_width"], 1)
        self.assertEqual(defaults["cen_dot_distance"], 37)
        self.assertEqual(defaults["biorientation_collinearity_threshold"], 3)
        self.assertEqual(defaults["puncta_line_mode"], "red_puncta")
        self.assertEqual(defaults["nuclear_cell_pair_mode"], "green_nucleus")
        self.assertEqual(defaults["nuclear_cell_pair_contour_mode"], "balanced")
        self.assertFalse(defaults["use_legacy_nuclear_cell_pair_pipeline"])
        self.assertEqual(defaults["green_dot_split_mode"], "balanced")
        self.assertFalse(defaults["red_dot_split_enabled"])
        self.assertEqual(defaults["red_dot_split_mode"], "balanced")
        self.assertNotIn("puncta_source_contour_count_filter", defaults)
        self.assertFalse(defaults["use_metadata_scale"])
        self.assertEqual(defaults["spatial_stats_unit"], "px")
        self.assertFalse(defaults["use_metadata_channel_order"])
        self.assertEqual(
            defaults["fallback_channel_order"],
            [CHANNEL_ROLE_GREEN, CHANNEL_ROLE_DIC, CHANNEL_ROLE_RED, CHANNEL_ROLE_BLUE],
        )
        self.assertFalse(normalized["auto_save_experiments"])
        self.assertFalse(normalized["confirm_cell_deletion"])
        self.assertFalse(normalized["confirm_multi_cell_deletion"])
        self.assertTrue(normalized["show_saved_file_channels"])
        self.assertFalse(normalized["show_saved_file_scales"])
        self.assertEqual(normalized["sidebar_spatial_stats_unit"], "px")
        self.assertEqual(normalized["main_image_channel"], "")
        self.assertEqual(
            normalized["default_puncta_source_contour_count_filter"],
            "all",
        )

    def test_normalize_preferences_migrates_legacy_green_split_default(self):
        normalized = normalize_preferences_payload(
            {
                "experiment_defaults": {
                    "biorientation_green_split_enabled": "false",
                },
            }
        )

        defaults = normalized["experiment_defaults"]
        self.assertFalse(defaults["green_dot_split_enabled"])
        self.assertNotIn("biorientation_green_split_enabled", defaults)

    def test_normalize_preferences_normalizes_result_display_default(self):
        cases = (
            (None, "all"),
            ("", "all"),
            ("bad_filter", "all"),
            ("all", "all"),
            ("exactly_1", "exactly_1"),
            ("exactly_2", "exactly_2"),
        )

        for raw_value, expected in cases:
            with self.subTest(raw_value=raw_value):
                normalized = normalize_preferences_payload(
                    {"default_puncta_source_contour_count_filter": raw_value}
                )

                self.assertEqual(
                    normalized["default_puncta_source_contour_count_filter"],
                    expected,
                )

    def test_legacy_experiment_contour_filter_does_not_create_display_default(self):
        normalized = normalize_preferences_payload(
            {"experiment_defaults": {"puncta_source_contour_count_filter": "2"}}
        )

        self.assertNotIn(
            "puncta_source_contour_count_filter",
            normalized["experiment_defaults"],
        )
        self.assertEqual(
            normalized["default_puncta_source_contour_count_filter"],
            "all",
        )

    def test_anonymous_preferences_default_result_display_filter_to_all(self):
        preferences = get_user_preferences(AnonymousUser())

        self.assertEqual(
            preferences["default_puncta_source_contour_count_filter"],
            "all",
        )

    def test_normalize_preferences_keeps_cen_dot_selection(self):
        normalized = normalize_preferences_payload(
            {"experiment_defaults": {"selected_plugins": ["CENDot"]}}
        )

        self.assertEqual(
            normalized["experiment_defaults"]["selected_plugins"],
            ["CENDot"],
        )
        self.assertFalse(
            normalized["experiment_defaults"]["signal_quantification_enabled"]
        )

    def test_normalize_preferences_migrates_legacy_primary_stats_to_puncta_mode(self):
        normalized = normalize_preferences_payload(
            {
                "experiment_defaults": {
                    "selected_plugins": [
                        "PunctaDistance",
                        "GreenRedIntensity",
                        "NuclearCellPairIntensity",
                        "CENDot",
                    ],
                }
            }
        )

        defaults = normalized["experiment_defaults"]
        self.assertEqual(
            defaults["selected_plugins"],
            ["PunctaDistance", "CENDot", "GreenRedIntensity"],
        )
        self.assertTrue(defaults["signal_quantification_enabled"])
        self.assertEqual(defaults["signal_quantification_mode"], "puncta_distance")
        self.assertTrue(defaults["puncta_contour_intensity_enabled"])

    def test_normalize_preferences_migrates_legacy_nuclear_only_mode(self):
        normalized = normalize_preferences_payload(
            {
                "experiment_defaults": {
                    "selected_plugins": [
                        "NuclearCellPairIntensity",
                        "Biorientation",
                    ],
                }
            }
        )

        defaults = normalized["experiment_defaults"]
        self.assertEqual(
            defaults["selected_plugins"],
            ["Biorientation", "NuclearCellPairIntensity"],
        )
        self.assertTrue(defaults["signal_quantification_enabled"])
        self.assertEqual(defaults["signal_quantification_mode"], "nuclear_cell_pair")
        self.assertFalse(defaults["puncta_contour_intensity_enabled"])

    def test_normalize_preferences_parent_off_keeps_independent_stats_only(self):
        normalized = normalize_preferences_payload(
            {
                "experiment_defaults": {
                    "selected_plugins": [
                        "PunctaDistance",
                        "GreenRedIntensity",
                        "CENDot",
                    ],
                    "signal_quantification_enabled": False,
                }
            }
        )

        defaults = normalized["experiment_defaults"]
        self.assertEqual(defaults["selected_plugins"], ["CENDot"])
        self.assertFalse(defaults["signal_quantification_enabled"])

    def test_normalize_preferences_accepts_legacy_alternate_detection_payload(self):
        normalized = normalize_preferences_payload(
            {
                "experiment_defaults": {
                    "selected_plugins": ["NuclearCellPairIntensity"],
                    "alternate_red_detection": True,
                }
            }
        )

        defaults = normalized["experiment_defaults"]
        self.assertTrue(defaults["alternate_nucleus_detection_enabled"])
        self.assertTrue(defaults["alternate_red_detection"])

    def test_signal_quantification_ignores_alternate_detection_in_puncta_mode(self):
        selection = resolve_signal_quantification_selection(
            payload={
                "signal_quantification_enabled": True,
                "signal_quantification_mode": "puncta_distance",
                "alternate_nucleus_detection_enabled": True,
            },
            selected_plugins=["PunctaDistance", "GreenRedIntensity"],
            nuclear_cell_pair_mode="green_nucleus",
        )

        self.assertTrue(selection.alternate_nucleus_detection_enabled)
        self.assertIsNone(selection.alternate_nucleus_detection_channel)

    def test_effective_alternate_detection_is_disabled_in_puncta_mode(self):
        enabled, channel = resolve_effective_alternate_nucleus_detection(
            signal_quantification_enabled=True,
            signal_quantification_mode="puncta_distance",
            nuclear_cell_pair_mode="red_nucleus",
            alternate_nucleus_detection_enabled=True,
            alternate_nucleus_detection_channel=CHANNEL_ROLE_GREEN,
        )

        self.assertFalse(enabled)
        self.assertIsNone(channel)

    def test_analysis_snapshot_disables_operational_alternate_detection_in_puncta_mode(
        self,
    ):
        normalized = normalize_analysis_config_snapshot(
            {
                "selected_analysis": ["PunctaDistance"],
                "signalQuantificationEnabled": True,
                "signalQuantificationMode": "puncta_distance",
                "alternateNucleusDetectionEnabled": True,
                "alternateNucleusDetectionChannel": CHANNEL_ROLE_GREEN,
                "nuclear_cell_pair_mode": "green_nucleus",
            }
        )

        self.assertFalse(normalized["alternateNucleusDetectionEnabled"])
        self.assertIsNone(normalized["alternateNucleusDetectionChannel"])
        self.assertFalse(normalized["alternateRedDetection"])

    def test_analysis_snapshot_preserves_legacy_scaled_nuclear_flag(self):
        normalized = normalize_analysis_config_snapshot(
            {
                "selected_analysis": ["NuclearCellPairIntensity"],
                "signalQuantificationEnabled": True,
                "signalQuantificationMode": "nuclear_cell_pair",
                "use_legacy_nuclear_cell_pair_pipeline": "true",
            }
        )

        self.assertTrue(normalized["use_legacy_nuclear_cell_pair_pipeline"])

    def test_analysis_snapshot_ignores_legacy_contour_count_filter_values(self):
        normalized = normalize_analysis_config_snapshot(
            {
                "selected_analysis": ["PunctaDistance"],
                "puncta_source_contour_count_filter": "exactly_2",
                "red_contour_count_filter": "exactly_1",
            }
        )

        self.assertNotIn("puncta_source_contour_count_filter", normalized)
        self.assertNotIn("red_contour_count_filter", normalized)

    def test_signal_quantification_applies_alternate_detection_in_nuclear_mode(self):
        selection = resolve_signal_quantification_selection(
            payload={
                "signal_quantification_enabled": True,
                "signal_quantification_mode": "nuclear_cell_pair",
                "alternate_nucleus_detection_enabled": True,
            },
            selected_plugins=["NuclearCellPairIntensity"],
            nuclear_cell_pair_mode="green_nucleus",
        )

        self.assertTrue(selection.alternate_nucleus_detection_enabled)
        self.assertEqual(
            selection.alternate_nucleus_detection_channel, CHANNEL_ROLE_GREEN
        )

    def test_effective_alternate_detection_derives_channel_in_nuclear_mode(self):
        enabled, channel = resolve_effective_alternate_nucleus_detection(
            signal_quantification_enabled=True,
            signal_quantification_mode="nuclear_cell_pair",
            nuclear_cell_pair_mode="red_nucleus",
            alternate_nucleus_detection_enabled=True,
            alternate_nucleus_detection_channel=None,
        )

        self.assertTrue(enabled)
        self.assertEqual(channel, CHANNEL_ROLE_RED)

    def test_signal_quantification_nuclear_mode_preserves_configured_secondary_plugins(
        self,
    ):
        selection = resolve_signal_quantification_selection(
            payload={
                "signal_quantification_enabled": True,
                "signal_quantification_mode": "nuclear_cell_pair",
                "puncta_contour_intensity_enabled": True,
            },
            selected_plugins=[
                "PunctaDistance",
                "GreenRedIntensity",
                "CENDot",
                "Biorientation",
            ],
            nuclear_cell_pair_mode="green_nucleus",
        )

        self.assertEqual(
            selection.configured_plugins,
            ("CENDot", "Biorientation", "NuclearCellPairIntensity"),
        )
        self.assertEqual(selection.selected_plugins, ("NuclearCellPairIntensity",))
        self.assertEqual(selection.paused_plugins, ("CENDot", "Biorientation"))
        self.assertFalse(selection.stat_visibility["cen_dot"])
        self.assertFalse(selection.stat_visibility["biorientation"])

    def test_sidebar_spatial_stats_unit_falls_back_to_workflow_default(self):
        normalized = normalize_preferences_payload(
            {
                "experiment_defaults": {"spatial_stats_unit": "um"},
            }
        )

        self.assertEqual(normalized["experiment_defaults"]["spatial_stats_unit"], "um")
        self.assertEqual(normalized["sidebar_spatial_stats_unit"], "um")

    def test_main_image_channel_accepts_supported_slug(self):
        normalized = normalize_preferences_payload({"main_image_channel": "green"})
        self.assertEqual(normalized["main_image_channel"], "green")

    def test_main_image_channel_drops_invalid_value(self):
        normalized = normalize_preferences_payload({"main_image_channel": "purple"})
        self.assertEqual(normalized["main_image_channel"], "")


class AccountAreaAccessTests(TestCase):
    def setUp(self):
        self.client = Client()
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            email="preference-tests@example.com",
            password="TestPass123!",
            first_name="Pref",
            last_name="Tester",
        )

    def test_account_area_requires_authentication(self):
        for name in ("dashboard", "account_settings", "workflow_defaults"):
            response = self.client.get(reverse(name))
            self.assertEqual(response.status_code, 302)
            self.assertIn(reverse("signin"), response["Location"])

    def test_delete_account_requires_matching_email(self):
        self.assertTrue(
            self.client.login(
                email="preference-tests@example.com",
                password="TestPass123!",
            )
        )
        response = self.client.post(
            reverse("account_settings"),
            {"action": "delete_account", "confirm_email": "wrong@example.com"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Incorrect email address entered.")
        self.assertTrue(get_user_model().objects.filter(pk=self.user.pk).exists())

    def test_account_settings_page_renders_information_and_actions_cards(self):
        self.assertTrue(
            self.client.login(
                email="preference-tests@example.com",
                password="TestPass123!",
            )
        )
        response = self.client.get(reverse("account_settings"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="accountInformationTitle"', html=False)
        self.assertContains(response, 'id="accountActionsTitle"', html=False)
        self.assertContains(response, "Account Information")
        self.assertContains(response, "Account Actions")
        self.assertContains(response, "Delete your account")

    def test_delete_account_removes_user_on_match(self):
        self.assertTrue(
            self.client.login(
                email="preference-tests@example.com",
                password="TestPass123!",
            )
        )
        response = self.client.post(
            reverse("account_settings"),
            {
                "action": "delete_account",
                "confirm_email": "preference-tests@example.com",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], reverse("home"))
        self.assertFalse(get_user_model().objects.filter(pk=self.user.pk).exists())


class AccountDeletionIntegrationTests(TestCase):
    def setUp(self):
        self.client = Client()
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            email="delete-owner@example.com",
            password="TestPass123!",
            first_name="Delete",
            last_name="Owner",
        )
        self.other_user = user_model.objects.create_user(
            email="delete-other@example.com",
            password="TestPass123!",
        )
        self.assertTrue(
            self.client.login(
                email="delete-owner@example.com",
                password="TestPass123!",
            )
        )

    def _create_account_artifacts(self, owner, stem: str) -> str:
        file_uuid = uuid4()
        uploaded = UploadedImage.objects.create(
            user=owner,
            name=stem,
            uuid=file_uuid,
            file_location=f"{file_uuid}/{stem}.dv",
        )
        segmented = SegmentedImage.objects.create(
            user=owner,
            UUID=file_uuid,
            file_location=f"user_{file_uuid}/{stem}.png",
            ImagePath=f"{file_uuid}/output/{stem}_frame_0.png",
            CellPairPrefix=f"{file_uuid}/segmented/cell_",
            NumCells=1,
        )
        CellStatistics.objects.create(
            segmented_image=segmented,
            cell_id=1,
            puncta_distance=1.0,
            puncta_line_intensity=2.0,
            nucleus_intensity_sum=3.0,
            cell_pair_intensity_sum=4.0,
        )
        self.assertTrue(UploadedImage.objects.filter(pk=uploaded.pk).exists())
        self.assertTrue(SegmentedImage.objects.filter(pk=segmented.pk).exists())
        self.assertTrue(
            CellStatistics.objects.filter(segmented_image=segmented).exists()
        )
        return str(file_uuid)

    def _create_media_artifacts(
        self, media_root: str, file_uuid: str, stem: str
    ) -> tuple[Path, Path]:
        uuid_dir = Path(media_root) / file_uuid
        user_uuid_dir = Path(media_root) / f"user_{file_uuid}"
        paths = [
            uuid_dir / f"{stem}.dv",
            uuid_dir / "output" / f"{stem}_frame_0.png",
            uuid_dir / "segmented" / "cell_1.png",
            user_uuid_dir / f"{stem}.png",
        ]
        for path in paths:
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(b"x")
        return uuid_dir, user_uuid_dir

    def test_delete_account_removes_user_related_rows_media_and_session(self):
        with TemporaryDirectory() as temp_media:
            owned_uuid = self._create_account_artifacts(self.user, "owned_sample")
            uuid_dir, user_uuid_dir = self._create_media_artifacts(
                temp_media,
                owned_uuid,
                "owned_sample",
            )
            self.assertTrue(uuid_dir.exists())
            self.assertTrue(user_uuid_dir.exists())

            with patch("accounts.views.profile.MEDIA_ROOT", temp_media):
                response = self.client.post(
                    reverse("account_settings"),
                    {
                        "action": "delete_account",
                        "confirm_email": "delete-owner@example.com",
                    },
                )

            self.assertEqual(response.status_code, 302)
            self.assertEqual(response["Location"], reverse("home"))
            self.assertFalse(get_user_model().objects.filter(pk=self.user.pk).exists())
            self.assertFalse(UploadedImage.objects.filter(uuid=owned_uuid).exists())
            self.assertFalse(SegmentedImage.objects.filter(UUID=owned_uuid).exists())
            self.assertFalse(
                CellStatistics.objects.filter(segmented_image_id=owned_uuid).exists()
            )
            self.assertFalse(uuid_dir.exists())
            self.assertFalse(user_uuid_dir.exists())

            auth_response = self.client.get(reverse("dashboard"))
            self.assertEqual(auth_response.status_code, 302)
            self.assertIn(reverse("signin"), auth_response["Location"])

    def test_delete_account_with_wrong_email_keeps_user_rows_and_media(self):
        with TemporaryDirectory() as temp_media:
            owned_uuid = self._create_account_artifacts(self.user, "keep_sample")
            uuid_dir, user_uuid_dir = self._create_media_artifacts(
                temp_media,
                owned_uuid,
                "keep_sample",
            )

            with patch("accounts.views.profile.MEDIA_ROOT", temp_media):
                response = self.client.post(
                    reverse("account_settings"),
                    {
                        "action": "delete_account",
                        "confirm_email": "wrong@example.com",
                    },
                )

            self.assertEqual(response.status_code, 200)
            self.assertContains(response, "Incorrect email address entered.")
            self.assertTrue(get_user_model().objects.filter(pk=self.user.pk).exists())
            self.assertTrue(UploadedImage.objects.filter(uuid=owned_uuid).exists())
            self.assertTrue(SegmentedImage.objects.filter(UUID=owned_uuid).exists())
            self.assertTrue(
                CellStatistics.objects.filter(segmented_image_id=owned_uuid).exists()
            )
            self.assertTrue(uuid_dir.exists())
            self.assertTrue(user_uuid_dir.exists())

    def test_delete_account_does_not_remove_other_users_data(self):
        with TemporaryDirectory() as temp_media:
            owned_uuid = self._create_account_artifacts(self.user, "owned_sample")
            other_uuid = self._create_account_artifacts(self.other_user, "other_sample")
            owned_uuid_dir, owned_user_dir = self._create_media_artifacts(
                temp_media,
                owned_uuid,
                "owned_sample",
            )
            other_uuid_dir, other_user_dir = self._create_media_artifacts(
                temp_media,
                other_uuid,
                "other_sample",
            )

            with patch("accounts.views.profile.MEDIA_ROOT", temp_media):
                response = self.client.post(
                    reverse("account_settings"),
                    {
                        "action": "delete_account",
                        "confirm_email": "delete-owner@example.com",
                    },
                )

            self.assertEqual(response.status_code, 302)
            self.assertFalse(get_user_model().objects.filter(pk=self.user.pk).exists())
            self.assertTrue(
                get_user_model().objects.filter(pk=self.other_user.pk).exists()
            )

            self.assertFalse(UploadedImage.objects.filter(uuid=owned_uuid).exists())
            self.assertFalse(SegmentedImage.objects.filter(UUID=owned_uuid).exists())
            self.assertFalse(
                CellStatistics.objects.filter(segmented_image_id=owned_uuid).exists()
            )
            self.assertFalse(owned_uuid_dir.exists())
            self.assertFalse(owned_user_dir.exists())

            self.assertTrue(UploadedImage.objects.filter(uuid=other_uuid).exists())
            self.assertTrue(SegmentedImage.objects.filter(UUID=other_uuid).exists())
            self.assertTrue(
                CellStatistics.objects.filter(segmented_image_id=other_uuid).exists()
            )
            self.assertTrue(other_uuid_dir.exists())
            self.assertTrue(other_user_dir.exists())


class DashboardBulkDeleteTests(TestCase):
    def setUp(self):
        self.client = Client()
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            email="dashboard-owner@example.com",
            password="TestPass123!",
        )
        self.other_user = user_model.objects.create_user(
            email="dashboard-other@example.com",
            password="TestPass123!",
        )
        self.assertTrue(
            self.client.login(
                email="dashboard-owner@example.com",
                password="TestPass123!",
            )
        )

    def _create_saved_file(self, owner, filename: str):
        file_uuid = uuid4()
        UploadedImage.objects.create(
            user=owner,
            name=filename,
            uuid=file_uuid,
            file_location=f"{file_uuid}/{filename}.dv",
        )
        SegmentedImage.objects.create(
            user=owner,
            UUID=file_uuid,
            file_location=f"user_{file_uuid}/{filename}.png",
            ImagePath=f"{file_uuid}/output/{filename}_frame_0.png",
            CellPairPrefix=f"{file_uuid}/segmented/cell_",
            NumCells=1,
        )
        return str(file_uuid)

    def test_bulk_delete_rejects_foreign_uuid(self):
        owned_uuid = self._create_saved_file(self.user, "owned")
        foreign_uuid = self._create_saved_file(self.other_user, "foreign")

        response = self.client.post(
            reverse("dashboard_bulk_delete"),
            data=json.dumps({"uuids": [owned_uuid, foreign_uuid]}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 403)
        self.assertTrue(UploadedImage.objects.filter(uuid=owned_uuid).exists())
        self.assertTrue(UploadedImage.objects.filter(uuid=foreign_uuid).exists())

    def test_bulk_delete_removes_owned_files(self):
        uuid_one = self._create_saved_file(self.user, "sample_one")
        uuid_two = self._create_saved_file(self.user, "sample_two")

        response = self.client.post(
            reverse("dashboard_bulk_delete"),
            data=json.dumps({"uuids": [uuid_one, uuid_two]}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["deleted_count"], 2)
        self.assertFalse(UploadedImage.objects.filter(uuid=uuid_one).exists())
        self.assertFalse(UploadedImage.objects.filter(uuid=uuid_two).exists())
        self.assertFalse(SegmentedImage.objects.filter(UUID=uuid_one).exists())
        self.assertFalse(SegmentedImage.objects.filter(UUID=uuid_two).exists())


class DisplayManualSaveTests(TestCase):
    def setUp(self):
        self.client = Client()
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            email="display-owner@example.com",
            password="TestPass123!",
        )
        self.other_user = user_model.objects.create_user(
            email="display-other@example.com",
            password="TestPass123!",
        )
        self.guest_user_id = get_guest_user()
        self.assertTrue(
            self.client.login(
                email="display-owner@example.com",
                password="TestPass123!",
            )
        )

    def _create_display_file(
        self,
        *,
        uploaded_owner,
        segmented_owner_id: str,
        filename: str,
    ) -> str:
        file_uuid = uuid4()
        UploadedImage.objects.create(
            user=uploaded_owner,
            name=filename,
            uuid=file_uuid,
            file_location=f"{file_uuid}/{filename}.dv",
        )
        SegmentedImage.objects.create(
            user_id=segmented_owner_id,
            UUID=file_uuid,
            file_location=f"user_{file_uuid}/{filename}.png",
            ImagePath=f"{file_uuid}/output/{filename}_frame_0.png",
            CellPairPrefix=f"{file_uuid}/segmented/cell_",
            NumCells=2,
        )
        return str(file_uuid)

    def _add_cell_stat(
        self,
        file_uuid: str,
        *,
        cell_id: int = 1,
        properties: dict | None = None,
    ) -> None:
        segmented = SegmentedImage.objects.get(UUID=file_uuid)
        stat_properties = {
            "signal_quantification_mode": "puncta_distance",
            "puncta_line_mode": "red_puncta",
            "nuclear_cell_pair_mode": "red_nucleus",
            "cen_dot_schema_version": 3,
            "puncta_distance_delta_x_px": 1.0,
            "puncta_distance_delta_y_px": 0.0,
            "distance_of_green_from_red_1_delta_x_px": 6.0,
            "distance_of_green_from_red_1_delta_y_px": 0.0,
            "red_contour_1_center_x_px": 10.0,
            "red_contour_1_center_y_px": 20.0,
            "green_contour_1_center_x_px": 30.0,
            "green_contour_1_center_y_px": 40.0,
            "blue_contour_center_x_px": 50.0,
            "blue_contour_center_y_px": 60.0,
        }
        if properties:
            stat_properties.update(properties)
        CellStatistics.objects.create(
            segmented_image=segmented,
            cell_id=cell_id,
            puncta_distance=1.0,
            puncta_line_intensity=2.0,
            nucleus_intensity_sum=3.0,
            cell_pair_intensity_sum=4.0,
            blue_contour_size=9.0,
            distance_of_green_from_red_1=6.0,
            red_in_red_total_intensity_1=5.0,
            red_in_red_max_intensity_1=4.0,
            red_in_red_average_intensity_1=2.5,
            green_in_red_total_intensity_1=6.0,
            green_in_red_max_intensity_1=5.0,
            green_in_red_average_intensity_1=3.0,
            red_in_green_total_intensity_1=7.0,
            red_in_green_max_intensity_1=6.0,
            red_in_green_average_intensity_1=3.5,
            green_in_green_total_intensity_1=8.0,
            green_in_green_max_intensity_1=7.0,
            green_in_green_average_intensity_1=4.0,
            green_red_intensity_1=6.0 / 5.0,
            category_cen_dot=1,
            properties=stat_properties,
        )

    def _set_transient_uuids(self, uuids: list[str]) -> None:
        session = self.client.session
        session["transient_experiment_uuids"] = uuids
        session.save()

    def _create_preprocess_file(self, *, filename: str) -> str:
        file_uuid = uuid4()
        uploaded = UploadedImage.objects.create(
            user=self.user,
            name=filename,
            uuid=file_uuid,
            file_location=f"{file_uuid}/{filename}.dv",
        )
        DVLayerTifPreview.objects.create(
            uploaded_image_uuid=uploaded,
            wavelength="DAPI",
            file_location=f"{file_uuid}/{filename}_preview.png",
        )
        return str(file_uuid)

    @staticmethod
    def _write_run_bytes(media_root: str, uuid_value: str, *, size: int) -> None:
        target = Path(media_root) / uuid_value / "output" / "frame.png"
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(b"x" * size)

    @staticmethod
    def _csv_rows(response) -> list[list[str]]:
        return list(csv.reader(StringIO(response.content.decode("utf-8"))))

    @staticmethod
    def _xlsx_headers(response) -> list[str]:
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        return [cell.value for cell in sheet[1]]

    @staticmethod
    def _xlsx_rows(response) -> list[list]:
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    @staticmethod
    def _fixed_export_time():
        return timezone.make_aware(
            datetime(2026, 5, 10, 18, 34),
            timezone.get_current_timezone(),
        )

    @staticmethod
    def _all_metric_columns() -> str:
        return ",".join(USER_SELECTABLE_TABLE_FIELDS)

    @staticmethod
    def _intensity_columns(
        *,
        statistics: tuple[str, ...] = ("total", "max", "average"),
        slots: tuple[int, ...] = (1, 2, 3),
        combinations: tuple[str, ...] = (
            "red_in_red",
            "green_in_red",
            "red_in_green",
            "green_in_green",
        ),
    ) -> list[str]:
        return [
            f"{combination}_{statistic}_intensity_{slot}"
            for combination in combinations
            for slot in slots
            for statistic in statistics
        ]

    def assertExportFilename(
        self,
        response,
        *,
        scope: str,
        file_count: int,
        extension: str,
    ) -> None:
        disposition = response["Content-Disposition"]
        self.assertIn("attachment;", disposition)
        self.assertRegex(
            disposition,
            (
                rf'filename="cytocv_{scope}_cell-metrics_{file_count}files_'
                rf'\d{{4}}-\d{{2}}-\d{{2}}_\d{{4}}\.{extension}"'
            ),
        )

    def assertExactExportFilename(
        self,
        response,
        *,
        scope: str,
        file_count: int,
        extension: str,
    ) -> None:
        self.assertEqual(
            response["Content-Disposition"],
            (
                f'attachment; filename="cytocv_{scope}_cell-metrics_'
                f'{file_count}files_2026-05-10_1834.{extension}"'
            ),
        )

    def test_display_save_endpoint_rejects_invalid_payload(self):
        response = self.client.post(
            reverse("display_save_files"),
            data=json.dumps({"uuids": "bad-shape"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 400)

    def test_display_save_endpoint_rejects_empty_uuid_list(self):
        response = self.client.post(
            reverse("display_save_files"),
            data=json.dumps({"uuids": []}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 400)

    def test_display_save_endpoint_rejects_foreign_or_unavailable_uuid(self):
        transient_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.guest_user_id,
            filename="transient_owned",
        )
        foreign_uuid = self._create_display_file(
            uploaded_owner=self.other_user,
            segmented_owner_id=self.guest_user_id,
            filename="foreign_uploaded",
        )
        self._set_transient_uuids([transient_uuid, foreign_uuid])

        response = self.client.post(
            reverse("display_save_files"),
            data=json.dumps({"uuids": [transient_uuid, foreign_uuid]}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(
            SegmentedImage.objects.get(UUID=transient_uuid).user_id,
            self.guest_user_id,
        )
        session = self.client.session
        self.assertIn(transient_uuid, session.get("transient_experiment_uuids", []))

    def test_display_save_endpoint_saves_transient_file_and_clears_session(self):
        transient_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.guest_user_id,
            filename="manual_save_candidate",
        )
        self._set_transient_uuids([transient_uuid])

        response = self.client.post(
            reverse("display_save_files"),
            data=json.dumps({"uuids": [transient_uuid]}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["saved_count"], 1)
        self.assertEqual(payload["already_saved_count"], 0)
        self.assertEqual(payload["saved_uuids"], [transient_uuid])
        self.assertEqual(
            SegmentedImage.objects.get(UUID=transient_uuid).user_id,
            self.user.id,
        )

        session = self.client.session
        self.assertNotIn(transient_uuid, session.get("transient_experiment_uuids", []))

        dashboard_response = self.client.get(reverse("dashboard"))
        self.assertEqual(dashboard_response.status_code, 200)
        self.assertContains(dashboard_response, "manual_save_candidate")

    def test_dashboard_renders_main_table_export_buttons(self):
        self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="dashboard_export_first",
        )

        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="celltable"', html=False)
        self.assertContains(response, 'id="exportButtons"', html=False)
        self.assertContains(response, 'id="downloadStatsBtn"', html=False)
        self.assertContains(response, "Download Statistics", html=False)
        self.assertNotContains(response, 'id="downloadCsvBtn"', html=False)
        self.assertNotContains(response, 'id="downloadXlsxBtn"', html=False)
        self.assertContains(response, 'id="downloadSelectedBtn"', html=False)
        self.assertContains(response, 'id="exportSelectionBackdrop"', html=False)
        self.assertContains(response, 'id="exportFileSelectionView"', html=False)
        self.assertContains(response, 'id="exportStatSelectionView"', html=False)
        self.assertContains(response, 'id="exportFormatToggle"', html=False)
        self.assertContains(response, 'data-active-format="csv"', html=False)
        self.assertContains(response, 'data-export-format="csv"', html=False)
        self.assertContains(response, 'data-export-format="xlsx"', html=False)
        self.assertContains(response, "Edit files", html=False)
        self.assertContains(response, 'id="exportSelectionConfig"', html=False)
        self.assertContains(response, "export_selection_modal.js", html=False)
        self.assertContains(
            response, 'id="deleteFilesStatus" aria-live="polite"', html=False
        )
        self.assertContains(
            response, '<span class="spinner" aria-hidden="true"></span>', html=False
        )
        self.assertContains(
            response, '<span class="btn-label">Confirm Delete</span>', html=False
        )
        self.assertContains(response, "js/pages/dashboard-file-actions.js", html=False)
        dashboard_actions_source = _frontend_static_text(
            "js/pages/dashboard-file-actions.js"
        )
        self.assertIn("let isDeletingFiles = false;", dashboard_actions_source)
        self.assertIn(
            "function setDeleteLoading(isLoading)", dashboard_actions_source
        )
        self.assertIn("Deleting selected files...", dashboard_actions_source)
        self.assertIn("if (isDeletingFiles) return;", dashboard_actions_source)
        self.assertNotContains(response, "sort=cell_id", html=False)
        self.assertNotContains(response, "data-file-export=", html=False)

    def test_dashboard_export_buttons_have_server_rendered_fallback_urls(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="dashboard_export_fallback",
        )
        self._add_cell_stat(saved_uuid)

        response = self.client.get(reverse("dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            f'href="/dashboard/?file_uuid={saved_uuid}&amp;_export=csv&amp;_unit=px"',
            html=False,
        )
        self.assertNotContains(response, "_export=xlsx", html=False)

    def test_results_payload_retains_all_rows_when_stale_contour_default_is_exact(self):
        self.user.config = {
            "preferences": {
                "experiment_defaults": {
                    "puncta_source_contour_count_filter": "exactly_2"
                }
            }
        }
        self.user.save(update_fields=["config"])
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="red_count_payload_retains_all",
        )
        self._add_cell_stat(
            saved_uuid,
            cell_id=1,
            properties={"puncta_source_contour_count": 1},
        )
        self._add_cell_stat(
            saved_uuid,
            cell_id=2,
            properties={"puncta_source_contour_count": 2},
        )

        cases = (
            (
                "dashboard",
                reverse("dashboard") + f"?file_uuid={saved_uuid}",
                "files_data_json",
            ),
            ("display", reverse("display", args=[saved_uuid]), "files_data"),
        )
        for route_name, url, context_key in cases:
            with self.subTest(route=route_name):
                response = self.client.get(url)

                self.assertEqual(response.status_code, 200)
                self.assertEqual(
                    response.context["puncta_source_contour_count_filter"],
                    "all",
                )
                files_data = json.loads(response.context[context_key])
                self.assertEqual(
                    sorted(files_data[saved_uuid]["Statistics"].keys()),
                    ["1", "2"],
                )

    def test_saved_result_display_default_initializes_dashboard_and_display(self):
        preferences = get_user_preferences(self.user)
        preferences["default_puncta_source_contour_count_filter"] = "exactly_1"
        update_user_preferences(self.user, preferences)
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="saved_source_filter_default",
        )
        self._add_cell_stat(
            saved_uuid,
            cell_id=1,
            properties={"puncta_source_contour_count": 1},
        )
        self._add_cell_stat(
            saved_uuid,
            cell_id=2,
            properties={"puncta_source_contour_count": 2},
        )

        cases = (
            (
                "dashboard",
                reverse("dashboard") + f"?file_uuid={saved_uuid}",
                "files_data_json",
                "dashboardPageConfig",
            ),
            (
                "display",
                reverse("display", args=[saved_uuid]),
                "files_data",
                "displayPageConfig",
            ),
        )
        for route_name, url, context_key, config_id in cases:
            with self.subTest(route=route_name):
                response = self.client.get(url)

                self.assertEqual(response.status_code, 200)
                self.assertEqual(
                    response.context["puncta_source_contour_count_filter"],
                    "exactly_1",
                )
                self.assertContains(
                    response,
                    f'<script id="{config_id}" type="application/json">',
                    html=False,
                )
                self.assertContains(
                    response,
                    '"initialPunctaSourceContourCountFilter": "exactly_1"',
                    html=False,
                )
                files_data = json.loads(response.context[context_key])
                self.assertEqual(
                    sorted(files_data[saved_uuid]["Statistics"].keys()),
                    ["1", "2"],
                )

    def test_invalid_result_display_default_initializes_to_all(self):
        self.user.config = {
            "preferences": {
                "default_puncta_source_contour_count_filter": "not-a-filter"
            }
        }
        self.user.save(update_fields=["config"])
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="invalid_source_filter_default",
        )
        self._add_cell_stat(saved_uuid)

        cases = (
            ("dashboard", reverse("dashboard") + f"?file_uuid={saved_uuid}"),
            ("display", reverse("display", args=[saved_uuid])),
        )
        for route_name, url in cases:
            with self.subTest(route=route_name):
                response = self.client.get(url)

                self.assertEqual(response.status_code, 200)
                self.assertEqual(
                    response.context["puncta_source_contour_count_filter"],
                    "all",
                )

    def test_query_contour_filter_takes_precedence_over_saved_display_default(self):
        preferences = get_user_preferences(self.user)
        preferences["default_puncta_source_contour_count_filter"] = "exactly_2"
        update_user_preferences(self.user, preferences)
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="query_source_filter_default",
        )
        self._add_cell_stat(saved_uuid)

        cases = (
            (
                "dashboard",
                reverse("dashboard")
                + f"?file_uuid={saved_uuid}&_puncta_source_contour_count=exactly_1",
            ),
            (
                "display",
                reverse("display", args=[saved_uuid])
                + "?_puncta_source_contour_count=exactly_1",
            ),
            (
                "display_legacy",
                reverse("display", args=[saved_uuid]) + "?_red_contour_count=exactly_1",
            ),
        )
        for route_name, url in cases:
            with self.subTest(route=route_name):
                response = self.client.get(url)

                self.assertEqual(response.status_code, 200)
                self.assertEqual(
                    response.context["puncta_source_contour_count_filter"],
                    "exactly_1",
                )

    def test_dashboard_template_renders_glass_layout_and_existing_hooks(self):
        self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="dashboard_glass_layout",
        )

        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response, 'data-ui-region="dashboard-main-shell"', html=False
        )
        self.assertContains(
            response, 'data-ui-region="dashboard-content-stack"', html=False
        )
        self.assertContains(response, 'data-ui-region="top-stage-card"', html=False)
        self.assertContains(response, 'data-ui-region="cell-pairs-card"', html=False)
        self.assertContains(response, 'data-ui-region="cell-metrics-strip"', html=False)
        self.assertContains(response, 'data-ui-region="stats-table-card"', html=False)
        self.assertContains(response, 'class="content-wrapper glass-shell"', html=False)
        self.assertContains(response, 'class="main-content glass-shell"', html=False)
        self.assertContains(
            response, 'class="storage-card glass-card glass-section"', html=False
        )
        self.assertContains(response, 'id="viewerPanel"', html=False)
        self.assertContains(response, 'id="mainChannelSwitcher"', html=False)
        self.assertContains(response, 'id="toggleContours"', html=False)
        self.assertContains(response, 'id="statsTablePanel"', html=False)
        self.assertContains(response, 'id="sidebarSpatialUnitToggle"', html=False)
        self.assertContains(response, 'id="tableFullscreenBtn"', html=False)
        self.assertContains(response, 'id="tableScrollFrame"', html=False)
        self.assertContains(response, 'id="downloadStatsBtn"', html=False)
        self.assertContains(response, 'data-action="select-cells"', html=False)
        self.assertContains(response, 'id="selectCellsBackdrop"', html=False)
        self.assertContains(response, 'id="dashboardPageConfig"', html=False)
        self.assertContains(response, "css/components/results-viewer.css", html=False)
        self.assertContains(response, "js/shared/results-viewer.js", html=False)
        self.assertContains(response, "js/shared/results-cell-actions.js", html=False)
        self.assertContains(response, "js/pages/dashboard-viewer.js", html=False)
        self.assertContains(response, "js/pages/dashboard-cell-actions.js", html=False)
        content = response.content.decode("utf-8")
        self.assertLess(
            content.index("css/components/results-viewer.css"),
            content.index("css/pages/dashboard.css"),
        )
        self.assertLess(
            content.index("js/shared/results-viewer.js"),
            content.index("js/pages/dashboard-viewer.js"),
        )
        self.assertLess(
            content.index("js/shared/results-cell-actions.js"),
            content.index("js/pages/dashboard-cell-actions.js"),
        )
        self.assertIn(
            "const initialSidebarSpatialStatsUnit =",
            _frontend_static_text("js/pages/dashboard-viewer.js"),
        )
        self.assertIn(
            "window.CytoCVResultsCellActions.init",
            _frontend_static_text("js/pages/dashboard-cell-actions.js"),
        )
        shared_cell_actions_source = _frontend_static_text(
            "js/shared/results-cell-actions.js"
        )
        self.assertIn("global.CytoCVResultsCellActions = { init };", shared_cell_actions_source)
        self.assertIn(
            "const tableFileUuid = (pageConfig && pageConfig.tableFileUuid) || '';",
            shared_cell_actions_source,
        )
        self.assertNotIn("CytoCVDashboardPageConfig", shared_cell_actions_source)
        self.assertNotIn("CytoCVDisplayPageConfig", shared_cell_actions_source)
        self.assertContains(
            response, 'id="previousFileBtn" disabled aria-disabled="true"', html=False
        )
        self.assertContains(
            response, 'id="nextFileBtn" disabled aria-disabled="true"', html=False
        )
        self.assertContains(response, "CEN Dot Measurements")
        self.assertContains(response, "Red In Red Total Intensity")
        self.assertContains(response, "Green In Red Total Intensity")
        self.assertContains(response, "Red In Green Total Intensity")
        self.assertContains(response, "Green In Green Total Intensity")
        self.assertContains(response, "Contour Intensities")
        self.assertContains(response, 'data-contour-intensity-display="total"', html=False)
        self.assertContains(response, 'data-contour-intensity-display="max"', html=False)
        self.assertContains(response, 'data-contour-intensity-display="average"', html=False)
        self.assertNotContains(response, "Intensity + Green Output")

    def test_display_template_renders_glass_layout_and_existing_hooks(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_glass_layout",
        )

        response = self.client.get(reverse("display", args=[saved_uuid]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-ui-region="display-main-shell"', html=False)
        self.assertContains(
            response, 'data-ui-region="display-content-stack"', html=False
        )
        self.assertContains(response, 'data-ui-region="top-stage-card"', html=False)
        self.assertContains(response, 'data-ui-region="cell-pairs-card"', html=False)
        self.assertContains(response, 'data-ui-region="cell-metrics-strip"', html=False)
        self.assertContains(response, 'data-ui-region="stats-table-card"', html=False)
        self.assertContains(response, 'class="content-wrapper glass-shell"', html=False)
        self.assertContains(response, 'class="main-content glass-shell"', html=False)
        self.assertContains(response, 'id="viewerPanel"', html=False)
        self.assertContains(response, 'id="mainChannelSwitcher"', html=False)
        self.assertContains(response, 'id="toggleContours"', html=False)
        self.assertContains(response, 'id="statsTablePanel"', html=False)
        self.assertContains(response, 'id="sidebarSpatialUnitToggle"', html=False)
        self.assertContains(response, 'id="tableFullscreenBtn"', html=False)
        self.assertContains(response, 'id="tableScrollFrame"', html=False)
        self.assertContains(response, 'id="celltable"', html=False)
        self.assertContains(response, 'id="displayExportButtons"', html=False)
        self.assertContains(response, 'id="displayDownloadStatsBtn"', html=False)
        self.assertContains(response, "Download Statistics", html=False)
        self.assertNotContains(response, 'id="displayDownloadCsvBtn"', html=False)
        self.assertNotContains(response, 'id="displayDownloadXlsxBtn"', html=False)
        self.assertContains(response, 'id="downloadSelectedBtn"', html=False)
        self.assertContains(response, 'id="exportSelectionBackdrop"', html=False)
        self.assertContains(response, 'id="exportFileSelectionView"', html=False)
        self.assertContains(response, 'id="exportStatSelectionView"', html=False)
        self.assertContains(response, 'id="exportFormatToggle"', html=False)
        self.assertContains(response, 'data-active-format="csv"', html=False)
        self.assertContains(response, 'data-export-format="csv"', html=False)
        self.assertContains(response, 'data-export-format="xlsx"', html=False)
        self.assertContains(response, "Edit files", html=False)
        self.assertContains(response, 'id="exportSelectionConfig"', html=False)
        self.assertContains(response, "export_selection_modal.js", html=False)
        self.assertContains(response, 'data-action="select-cells"', html=False)
        self.assertContains(response, 'id="selectCellsBackdrop"', html=False)
        self.assertNotContains(response, "sort=cell_id", html=False)
        self.assertContains(response, 'id="displayPageConfig"', html=False)
        self.assertContains(response, "css/components/results-viewer.css", html=False)
        self.assertContains(response, "js/shared/results-viewer.js", html=False)
        self.assertContains(response, "js/shared/results-cell-actions.js", html=False)
        self.assertContains(response, "js/pages/display-viewer.js", html=False)
        self.assertContains(response, "js/pages/display-cell-actions.js", html=False)
        content = response.content.decode("utf-8")
        self.assertLess(
            content.index("css/components/results-viewer.css"),
            content.index("css/pages/display.css"),
        )
        self.assertLess(
            content.index("js/shared/results-viewer.js"),
            content.index("js/pages/display-viewer.js"),
        )
        self.assertLess(
            content.index("js/shared/results-cell-actions.js"),
            content.index("js/pages/display-cell-actions.js"),
        )
        display_viewer_source = _frontend_static_text("js/pages/display-viewer.js")
        self.assertIn("const defaultSpatialStatsUnit =", display_viewer_source)
        self.assertIn(
            "const initialSidebarSpatialStatsUnit =",
            display_viewer_source,
        )
        self.assertIn(
            "window.CytoCVResultsCellActions.init",
            _frontend_static_text("js/pages/display-cell-actions.js"),
        )
        self.assertContains(
            response, 'id="previousFileBtn" disabled aria-disabled="true"', html=False
        )
        self.assertContains(
            response, 'id="nextFileBtn" disabled aria-disabled="true"', html=False
        )
        self.assertContains(response, 'id="dic_form"', html=False)
        self.assertContains(response, 'id="blue_form"', html=False)
        self.assertContains(response, 'id="red_form"', html=False)
        self.assertContains(response, 'id="green_form"', html=False)
        self.assertContains(response, "CEN Dot Measurements")
        self.assertContains(response, "Red In Red Total Intensity")
        self.assertContains(response, "Green In Red Total Intensity")
        self.assertContains(response, "Red In Green Total Intensity")
        self.assertContains(response, "Green In Green Total Intensity")
        self.assertContains(response, "Contour Intensities")
        self.assertContains(response, 'data-contour-intensity-display="total"', html=False)
        self.assertContains(response, 'data-contour-intensity-display="max"', html=False)
        self.assertContains(response, 'data-contour-intensity-display="average"', html=False)
        self.assertNotContains(response, "Intensity + Green Output")

    def test_display_export_buttons_are_not_bound_to_initial_table_uuid(self):
        first_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_export_first",
        )
        second_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_export_second",
        )
        self._add_cell_stat(first_uuid)
        self._add_cell_stat(second_uuid)

        response = self.client.get(
            reverse("display", args=[f"{first_uuid},{second_uuid}"])
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "js/pages/display-viewer.js", html=False)
        display_viewer_source = _frontend_static_text("js/pages/display-viewer.js")
        self.assertIn("function syncDisplayExportButtons", display_viewer_source)
        self.assertIn(
            "syncDisplayExportButtons(fileUUID, fileData, renderedRowCount);",
            display_viewer_source,
        )
        self.assertNotIn("serverTableUUID", display_viewer_source)

    def test_display_view_serializes_nuclear_contour_source_without_stat_card_row(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_contour_source",
        )
        self._add_cell_stat(saved_uuid)
        cell_stat = CellStatistics.objects.get(
            segmented_image__UUID=saved_uuid,
            cell_id=1,
        )
        properties = dict(cell_stat.properties or {})
        properties.update(
            {
                "selected_analysis": ["NuclearCellPairIntensity"],
                "nuclear_cell_pair_contour_channel": "Red",
                "nuclear_cell_pair_measurement_channel": "Green",
                "nuclear_cell_pair_contour_source": "alternate_red_nucleus_slot_1",
                "nuclear_cell_pair_status": "ok",
            }
        )
        cell_stat.properties = properties
        cell_stat.save(update_fields=["properties"])

        response = self.client.get(reverse("display", args=[saved_uuid]))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'id="nucleusContourSource"', html=False)
        self.assertNotContains(response, "Contour Source Used", html=False)
        files_data = json.loads(response.context["files_data"])
        payload = files_data[saved_uuid]["Statistics"]["1"]
        self.assertEqual(
            payload["nuclear_cell_pair_contour_source"],
            "alternate_red_nucleus_slot_1",
        )

    def test_dashboard_template_exposes_preferred_main_image_channel(self):
        self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="dashboard_preferred_main_channel",
        )
        preferences = get_user_preferences(self.user)
        preferences["main_image_channel"] = "green"
        update_user_preferences(self.user, preferences)

        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            '"initialPreferredMainImageChannel": "green"',
            html=False,
        )

    def test_display_template_exposes_preferred_main_image_channel(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_preferred_main_channel",
        )
        preferences = get_user_preferences(self.user)
        preferences["main_image_channel"] = "green"
        update_user_preferences(self.user, preferences)

        response = self.client.get(reverse("display", args=[saved_uuid]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            '"initialPreferredMainImageChannel": "green"',
            html=False,
        )

    def test_dashboard_and_display_templates_expose_cell_delete_confirmation_preference(
        self,
    ):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="cell_delete_confirmation_pref",
        )
        preferences = get_user_preferences(self.user)
        preferences["confirm_cell_deletion"] = False
        preferences["confirm_multi_cell_deletion"] = False
        update_user_preferences(self.user, preferences)

        dashboard_response = self.client.get(reverse("dashboard"))
        display_response = self.client.get(reverse("display", args=[saved_uuid]))

        self.assertEqual(dashboard_response.status_code, 200)
        self.assertEqual(display_response.status_code, 200)
        self.assertContains(
            dashboard_response,
            '"confirmCellDeletion": false',
            html=False,
        )
        self.assertContains(
            dashboard_response,
            '"confirmMultiCellDeletion": false',
            html=False,
        )
        self.assertContains(
            display_response,
            '"confirmCellDeletion": false',
            html=False,
        )
        self.assertContains(
            display_response,
            '"confirmMultiCellDeletion": false',
            html=False,
        )

    def test_preprocess_template_renders_glass_layout_and_existing_hooks(self):
        preprocess_uuid = self._create_preprocess_file(
            filename="preprocess_glass_layout"
        )

        response = self.client.get(reverse("pre_process", args=[preprocess_uuid]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response, 'data-ui-region="preprocess-main-shell"', html=False
        )
        self.assertContains(
            response, 'data-ui-region="preprocess-content-stack"', html=False
        )
        self.assertContains(response, 'data-ui-region="file-context-card"', html=False)
        self.assertContains(response, 'data-ui-region="main-image-stage"', html=False)
        self.assertNotContains(response, 'data-ui-region="actions-card"', html=False)
        self.assertContains(response, 'class="content-wrapper glass-shell"', html=False)
        self.assertContains(response, 'class="main-content glass-shell"', html=False)
        self.assertContains(response, 'id="preprocessForm"', html=False)
        self.assertContains(response, 'id="imageContainer"', html=False)
        self.assertContains(response, 'id="prevButton"', html=False)
        self.assertContains(response, 'id="nextButton"', html=False)
        self.assertContains(response, 'id="currentFileInfo"', html=False)
        self.assertContains(response, 'id="currentFileIndex"', html=False)
        self.assertContains(response, 'id="preprocessScaleSummary"', html=False)
        self.assertContains(response, 'id="sidebarSpatialUnitToggle"', html=False)

    def test_dashboard_csv_export_for_file_uuid_returns_named_attachment(self):
        file_name = "dashboard_csv_export"
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename=file_name,
        )
        self._add_cell_stat(saved_uuid)

        response = self.client.get(
            reverse("dashboard"),
            {"file_uuid": saved_uuid, "_export": "csv"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertExportFilename(
            response,
            scope="all",
            file_count=1,
            extension="csv",
        )
        self.assertIn("text/csv", response["Content-Type"])
        csv_text = response.content.decode("utf-8")
        self.assertIn("Cell ID", csv_text)
        self.assertIn("Red In Red Total Intensity 1", csv_text)
        self.assertIn("Red In Red Max Intensity 1", csv_text)
        self.assertIn("Red In Red Average Intensity 1", csv_text)
        self.assertIn("Green In Red Total Intensity 1", csv_text)
        self.assertIn("Green In Red Max Intensity 1", csv_text)
        self.assertIn("Green In Red Average Intensity 1", csv_text)
        self.assertIn("Red In Green Total Intensity 1", csv_text)
        self.assertIn("Red In Green Max Intensity 1", csv_text)
        self.assertIn("Red In Green Average Intensity 1", csv_text)
        self.assertIn("Green In Green Total Intensity 1", csv_text)
        self.assertIn("Green In Green Max Intensity 1", csv_text)
        self.assertIn("Green In Green Average Intensity 1", csv_text)
        self.assertNotIn("Green/Red ratio 1", csv_text)
        self.assertIn("Measurement/Contour Ratio 1 (Green/Red)", csv_text)
        self.assertIn("5.000", csv_text)
        self.assertIn("8.000", csv_text)
        self.assertIn("Cen Dot Location", csv_text)
        self.assertIn("Mother and daughter", csv_text)

    def test_dashboard_xlsx_export_for_file_uuid_returns_named_attachment(self):
        file_name = "dashboard_xlsx_export"
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename=file_name,
        )
        self._add_cell_stat(saved_uuid)

        response = self.client.get(
            reverse("dashboard"),
            {"file_uuid": saved_uuid, "_export": "xlsx"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertExportFilename(
            response,
            scope="all",
            file_count=1,
            extension="xlsx",
        )
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )
        self.assertGreater(len(response.content), 0)
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        self.assertIn("Red In Red Total Intensity 1", headers)
        self.assertIn("Red In Red Max Intensity 1", headers)
        self.assertIn("Red In Red Average Intensity 1", headers)
        self.assertIn("Green In Red Total Intensity 1", headers)
        self.assertIn("Green In Red Max Intensity 1", headers)
        self.assertIn("Green In Red Average Intensity 1", headers)
        self.assertIn("Red In Green Total Intensity 1", headers)
        self.assertIn("Red In Green Max Intensity 1", headers)
        self.assertIn("Red In Green Average Intensity 1", headers)
        self.assertIn("Green In Green Total Intensity 1", headers)
        self.assertIn("Green In Green Max Intensity 1", headers)
        self.assertIn("Green In Green Average Intensity 1", headers)
        self.assertNotIn("Green/Red ratio 1", headers)
        self.assertIn("Measurement/Contour Ratio 1 (Green/Red)", headers)
        self.assertEqual(sheet.cell(row=2, column=1).value, 1)
        self.assertEqual(sheet.cell(row=2, column=1).data_type, "n")
        red_intensity_col = headers.index("Red In Red Total Intensity 1") + 1
        red_intensity_cell = sheet.cell(row=2, column=red_intensity_col)
        self.assertEqual(red_intensity_cell.value, 5)
        self.assertEqual(red_intensity_cell.data_type, "n")
        cen_dot_col = headers.index("Cen Dot Location") + 1
        cen_dot_cell = sheet.cell(row=2, column=cen_dot_col)
        self.assertEqual(cen_dot_cell.value, "Mother and daughter")
        self.assertEqual(cen_dot_cell.data_type, "s")

    def test_dashboard_single_export_filename_scope_tracks_metric_selection(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="dashboard_filename_metric_scope",
        )
        other_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="dashboard_filename_other_file",
        )
        self._add_cell_stat(saved_uuid)
        self._add_cell_stat(other_uuid)

        cases = [
            ("csv", {}, "all", "csv"),
            ("xlsx", {}, "all", "xlsx"),
            (
                "csv",
                {"_columns": self._all_metric_columns()},
                "all",
                "csv",
            ),
            (
                "xlsx",
                {"_columns": self._all_metric_columns()},
                "all",
                "xlsx",
            ),
            (
                "csv",
                {"_columns": "red_in_red_total_intensity_1,puncta_distance"},
                "selected",
                "csv",
            ),
            (
                "xlsx",
                {"_columns": "red_in_red_total_intensity_1,puncta_distance"},
                "selected",
                "xlsx",
            ),
        ]
        with patch(
            "core.services.export_filenames.timezone.now",
            return_value=self._fixed_export_time(),
        ):
            for export_format, extra_params, expected_scope, extension in cases:
                with self.subTest(export_format=export_format, params=extra_params):
                    response = self.client.get(
                        reverse("dashboard"),
                        {
                            "file_uuid": saved_uuid,
                            "_export": export_format,
                            **extra_params,
                        },
                    )

                    self.assertEqual(response.status_code, 200)
                    self.assertExactExportFilename(
                        response,
                        scope=expected_scope,
                        file_count=1,
                        extension=extension,
                    )

    def test_dashboard_csv_export_filters_selected_columns(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="dashboard_filtered_csv",
        )
        self._add_cell_stat(saved_uuid)

        response = self.client.get(
            reverse("dashboard"),
            {
                "file_uuid": saved_uuid,
                "_export": "csv",
                "_columns": (
                    "cytoplasmic_intensity,red_in_red_total_intensity_1,"
                    "red_contour_1_center_xy,puncta_distance"
                ),
            },
        )

        self.assertEqual(response.status_code, 200)
        rows = self._csv_rows(response)
        self.assertEqual(
            rows[0],
            [
                "Cell ID",
                "Distance Between Red Puncta (px)",
                "Red Contour 1 Center (x,y) (px)",
                "Red In Red Total Intensity 1",
                "Cytoplasmic Intensity",
            ],
        )
        self.assertEqual(rows[1], ["1", "1.000", "10.000, 20.000", "5.000", "0.000"])

    def test_dashboard_xlsx_export_filters_selected_columns(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="dashboard_filtered_xlsx",
        )
        self._add_cell_stat(saved_uuid)

        response = self.client.get(
            reverse("dashboard"),
            {
                "file_uuid": saved_uuid,
                "_export": "xlsx",
                "_columns": "red_in_red_total_intensity_1,measurement_contour_ratio_1",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            self._xlsx_headers(response),
            [
                "Cell ID",
                "Red In Red Total Intensity 1",
                "Measurement/Contour Ratio 1 (Green/Red)",
            ],
        )

    def test_single_selected_exports_respect_red_count_filter_without_changing_columns(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="single_red_count_filter",
        )
        self._add_cell_stat(saved_uuid, cell_id=1, properties={"puncta_source_contour_count": 1})
        self._add_cell_stat(saved_uuid, cell_id=2, properties={"puncta_source_contour_count": 2})
        selected_columns = "puncta_distance,red_in_red_total_intensity_1"
        expected_headers = [
            "Cell ID",
            "Distance Between Red Puncta (px)",
            "Red In Red Total Intensity 1",
        ]

        request_cases = (
            ("dashboard", reverse("dashboard"), {"file_uuid": saved_uuid}),
            ("display", reverse("display", args=[saved_uuid]), {}),
        )
        for route_name, url, base_params in request_cases:
            for export_format in ("csv", "xlsx"):
                with self.subTest(route=route_name, export_format=export_format):
                    response = self.client.get(
                        url,
                        {
                            **base_params,
                            "_export": export_format,
                            "_columns": selected_columns,
                            "_puncta_source_contour_count": "exactly_2",
                        },
                    )

                    self.assertEqual(response.status_code, 200)
                    if export_format == "csv":
                        rows = self._csv_rows(response)
                        self.assertEqual(rows[0], expected_headers)
                        self.assertEqual(rows[1:], [["2", "1.000", "5.000"]])
                    else:
                        rows = self._xlsx_rows(response)
                        self.assertEqual(rows[0], expected_headers)
                        self.assertEqual(rows[1:], [[2, 1, 5]])

                    all_response = self.client.get(
                        url,
                        {
                            **base_params,
                            "_export": export_format,
                            "_columns": selected_columns,
                            "_puncta_source_contour_count": "invalid",
                        },
                    )
                    self.assertEqual(all_response.status_code, 200)
                    if export_format == "csv":
                        all_rows = self._csv_rows(all_response)
                        self.assertEqual(all_rows[0], expected_headers)
                        self.assertEqual(
                            [row[0] for row in all_rows[1:]],
                            ["1", "2"],
                        )
                    else:
                        all_rows = self._xlsx_rows(all_response)
                        self.assertEqual(all_rows[0], expected_headers)
                        self.assertEqual(
                            [row[0] for row in all_rows[1:]],
                            [1, 2],
                        )

    def test_dashboard_selected_intensity_exports_keep_total_max_and_average_independent(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="dashboard_selected_intensity_independent",
        )
        self._add_cell_stat(saved_uuid)

        total_only_fields = (
            "red_in_red_total_intensity_1,"
            "green_in_red_total_intensity_1,"
            "red_in_green_total_intensity_1,"
            "green_in_green_total_intensity_1"
        )
        response = self.client.get(
            reverse("dashboard"),
            {
                "file_uuid": saved_uuid,
                "_export": "csv",
                "_columns": total_only_fields,
            },
        )
        rows = self._csv_rows(response)
        self.assertEqual(
            rows[0],
            [
                "Cell ID",
                "Red In Red Total Intensity 1",
                "Green In Red Total Intensity 1",
                "Red In Green Total Intensity 1",
                "Green In Green Total Intensity 1",
            ],
        )
        self.assertEqual(rows[1], ["1", "5.000", "6.000", "7.000", "8.000"])

        response = self.client.get(
            reverse("dashboard"),
            {
                "file_uuid": saved_uuid,
                "_export": "xlsx",
                "_columns": total_only_fields,
            },
        )
        self.assertEqual(
            self._xlsx_headers(response),
            [
                "Cell ID",
                "Red In Red Total Intensity 1",
                "Green In Red Total Intensity 1",
                "Red In Green Total Intensity 1",
                "Green In Green Total Intensity 1",
            ],
        )
        self.assertEqual(self._xlsx_rows(response)[1], [1, 5, 6, 7, 8])

        response = self.client.get(
            reverse("dashboard"),
            {
                "file_uuid": saved_uuid,
                "_export": "csv",
                "_columns": (
                    "red_in_red_total_intensity_1,"
                    "red_in_red_max_intensity_1,"
                    "green_in_red_total_intensity_1,"
                    "green_in_red_max_intensity_1"
                ),
            },
        )
        rows = self._csv_rows(response)
        self.assertEqual(
            rows[0],
            [
                "Cell ID",
                "Red In Red Total Intensity 1",
                "Red In Red Max Intensity 1",
                "Green In Red Total Intensity 1",
                "Green In Red Max Intensity 1",
            ],
        )
        self.assertNotIn("Red In Red Average Intensity 1", rows[0])
        self.assertEqual(rows[1], ["1", "5.000", "4.000", "6.000", "5.000"])

        response = self.client.get(
            reverse("dashboard"),
            {
                "file_uuid": saved_uuid,
                "_export": "xlsx",
                "_columns": (
                    "red_in_red_average_intensity_1,"
                    "green_in_red_average_intensity_1,"
                    "red_in_green_average_intensity_1,"
                    "green_in_green_average_intensity_1"
                ),
            },
        )
        self.assertEqual(
            self._xlsx_headers(response),
            [
                "Cell ID",
                "Red In Red Average Intensity 1",
                "Green In Red Average Intensity 1",
                "Red In Green Average Intensity 1",
                "Green In Green Average Intensity 1",
            ],
        )
        self.assertEqual(self._xlsx_rows(response)[1], [1, 2.5, 3, 3.5, 4])

    def test_display_selected_total_only_intensity_exports_all_slots_and_combinations(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_selected_total_only_intensity",
        )
        self._add_cell_stat(saved_uuid)
        columns = ",".join(self._intensity_columns(statistics=("total",)))

        csv_response = self.client.get(
            reverse("display", args=[saved_uuid]),
            {"_export": "csv", "_columns": columns},
        )
        xlsx_response = self.client.get(
            reverse("display", args=[saved_uuid]),
            {"_export": "xlsx", "_columns": columns},
        )

        self.assertEqual(csv_response.status_code, 200)
        self.assertEqual(xlsx_response.status_code, 200)
        csv_headers = self._csv_rows(csv_response)[0]
        xlsx_headers = self._xlsx_headers(xlsx_response)
        self.assertEqual(csv_headers, xlsx_headers)
        self.assertEqual(len(csv_headers), 13)
        self.assertIn("Red In Red Total Intensity 1", csv_headers)
        self.assertIn("Green In Green Total Intensity 3", csv_headers)
        self.assertTrue(all("Total Intensity" in header for header in csv_headers[1:]))
        self.assertFalse(any("Max Intensity" in header for header in csv_headers))
        self.assertFalse(any("Average Intensity" in header for header in csv_headers))

    def test_dashboard_selected_slots_one_two_intensity_exports_exclude_slot_three(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="dashboard_selected_slots_one_two_intensity",
        )
        self._add_cell_stat(saved_uuid)
        columns = ",".join(self._intensity_columns(slots=(1, 2)))

        for export_format in ("csv", "xlsx"):
            with self.subTest(export_format=export_format):
                response = self.client.get(
                    reverse("dashboard"),
                    {
                        "file_uuid": saved_uuid,
                        "_export": export_format,
                        "_columns": columns,
                    },
                )
                self.assertEqual(response.status_code, 200)
                headers = (
                    self._csv_rows(response)[0]
                    if export_format == "csv"
                    else self._xlsx_headers(response)
                )
                self.assertEqual(len(headers), 25)
                self.assertIn("Red In Red Total Intensity 1", headers)
                self.assertIn("Green In Green Average Intensity 2", headers)
                self.assertFalse(any(header.endswith("Intensity 3") for header in headers))

    def test_display_selected_same_channel_total_intensity_slots_one_two(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_selected_same_channel_intensity",
        )
        self._add_cell_stat(saved_uuid)
        columns = ",".join(
            self._intensity_columns(
                statistics=("total",),
                slots=(1, 2),
                combinations=("red_in_red", "green_in_green"),
            )
        )

        response = self.client.get(
            reverse("display", args=[saved_uuid]),
            {"_export": "csv", "_columns": columns},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            self._csv_rows(response)[0],
            [
                "Cell ID",
                "Red In Red Total Intensity 1",
                "Red In Red Total Intensity 2",
                "Green In Green Total Intensity 1",
                "Green In Green Total Intensity 2",
            ],
        )

    def test_dashboard_csv_export_respects_micron_unit_request(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="dashboard_csv_export_um",
        )
        uploaded = UploadedImage.objects.get(uuid=saved_uuid)
        uploaded.scale_info = build_scale_info(
            manual_um_per_px=0.5, prefer_metadata=False
        )
        uploaded.save(update_fields=["scale_info"])
        self._add_cell_stat(saved_uuid)

        response = self.client.get(
            reverse("dashboard"),
            {"file_uuid": saved_uuid, "_export": "csv", "_unit": "um"},
        )

        self.assertEqual(response.status_code, 200)
        csv_text = response.content.decode("utf-8")
        self.assertIn("Distance Between Red Puncta (µm)", csv_text)
        self.assertIn("Blue Contour Size (µm²)", csv_text)
        self.assertIn("Distance Of Green From Red 1 (µm)", csv_text)
        self.assertIn("0.500", csv_text)
        self.assertIn("2.250", csv_text)
        self.assertIn("3.000", csv_text)

    def test_filtered_dashboard_csv_export_preserves_order_and_spatial_unit(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="dashboard_filtered_export_um",
        )
        uploaded = UploadedImage.objects.get(uuid=saved_uuid)
        uploaded.scale_info = build_scale_info(
            manual_um_per_px=0.5, prefer_metadata=False
        )
        uploaded.save(update_fields=["scale_info"])
        self._add_cell_stat(saved_uuid)

        response = self.client.get(
            reverse("dashboard"),
            {
                "file_uuid": saved_uuid,
                "_export": "csv",
                "_unit": "um",
                "_columns": "red_in_red_total_intensity_1,puncta_distance",
            },
        )

        self.assertEqual(response.status_code, 200)
        rows = self._csv_rows(response)
        self.assertEqual(
            rows[0],
            ["Cell ID", "Distance Between Red Puncta (µm)", "Red In Red Total Intensity 1"],
        )
        self.assertEqual(rows[1], ["1", "0.500", "5.000"])

    def test_dashboard_xlsx_export_respects_micron_unit_request(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="dashboard_xlsx_export_um",
        )
        uploaded = UploadedImage.objects.get(uuid=saved_uuid)
        uploaded.scale_info = build_scale_info(
            manual_um_per_px=0.5, prefer_metadata=False
        )
        uploaded.save(update_fields=["scale_info"])
        self._add_cell_stat(saved_uuid)

        response = self.client.get(
            reverse("dashboard"),
            {"file_uuid": saved_uuid, "_export": "xlsx", "_unit": "um"},
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        self.assertIn("Distance Between Red Puncta (µm)", headers)
        self.assertIn("Blue Contour Size (µm²)", headers)
        self.assertIn("Distance Of Green From Red 1 (µm)", headers)
        self.assertIn("Cen Dot Location", headers)
        gfp_dot_col = headers.index("Cen Dot Location") + 1
        self.assertEqual(
            sheet.cell(row=2, column=gfp_dot_col).value, "Mother and daughter"
        )

    def test_dashboard_combined_csv_export_filters_stats_and_preserves_file_order(self):
        first_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="combined_first",
        )
        second_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="combined_second",
        )
        self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="combined_not_selected",
        )
        self._add_cell_stat(first_uuid, cell_id=2)
        self._add_cell_stat(first_uuid, cell_id=1)
        self._add_cell_stat(second_uuid, cell_id=1)

        response = self.client.post(
            reverse("dashboard_bulk_export"),
            data=json.dumps(
                {
                    "uuids": [second_uuid, first_uuid],
                    "_export": "csv",
                    "_columns": [
                        "red_in_red_total_intensity_1",
                        "puncta_distance",
                        "red_contour_1_center_xy",
                    ],
                    "_unit": "px",
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertExportFilename(
            response,
            scope="selected",
            file_count=2,
            extension="csv",
        )
        rows = self._csv_rows(response)
        self.assertEqual(
            rows[0],
            [
                "File Name",
                "Cell ID",
                "Puncta Distance (px)",
                "Red Contour 1 Center (x,y) (px)",
                "Red In Red Total Intensity 1",
            ],
        )
        self.assertNotIn("Red In Red Max Intensity 1", rows[0])
        self.assertNotIn("Red In Red Average Intensity 1", rows[0])
        self.assertEqual(
            rows[1:],
            [
                ["combined_second", "1", "1.000", "10.000, 20.000", "5.000"],
                ["combined_first", "1", "1.000", "10.000, 20.000", "5.000"],
                ["", "2", "1.000", "10.000, 20.000", "5.000"],
            ],
        )

    def test_dashboard_combined_xlsx_export_matches_filtered_headers_and_order(self):
        first_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="combined_xlsx_first",
        )
        second_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="combined_xlsx_second",
        )
        self._add_cell_stat(first_uuid)
        self._add_cell_stat(second_uuid)

        response = self.client.post(
            reverse("dashboard_bulk_export"),
            data=json.dumps(
                {
                    "uuids": [first_uuid, second_uuid],
                    "_export": "xlsx",
                    "_columns": ["red_in_red_total_intensity_1", "measurement_contour_ratio_1"],
                    "_unit": "px",
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertExportFilename(
            response,
            scope="selected",
            file_count=2,
            extension="xlsx",
        )
        rows = self._xlsx_rows(response)
        self.assertEqual(
            rows[0],
            [
                "File Name",
                "Cell ID",
                "Red In Red Total Intensity 1",
                "Measurement/Contour Ratio 1",
            ],
        )
        self.assertNotIn("Red In Red Max Intensity 1", rows[0])
        self.assertNotIn("Red In Red Average Intensity 1", rows[0])
        self.assertEqual(
            [row[0] for row in rows[1:]],
            ["combined_xlsx_first", "combined_xlsx_second"],
        )
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        red_intensity_cell = sheet.cell(row=2, column=3)
        ratio_cell = sheet.cell(row=2, column=4)
        self.assertEqual(red_intensity_cell.value, 5)
        self.assertEqual(red_intensity_cell.data_type, "n")
        self.assertEqual(ratio_cell.data_type, "n")

    def test_combined_exports_respect_red_count_filter_without_changing_columns(self):
        first_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="combined_red_filter_first",
        )
        second_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="combined_red_filter_second",
        )
        self._add_cell_stat(first_uuid, cell_id=1, properties={"puncta_source_contour_count": 1})
        self._add_cell_stat(first_uuid, cell_id=2, properties={"puncta_source_contour_count": 2})
        self._add_cell_stat(second_uuid, cell_id=1, properties={"puncta_source_contour_count": 1})
        columns = ["red_in_red_total_intensity_1", "puncta_distance"]
        expected_headers = [
            "File Name",
            "Cell ID",
            "Puncta Distance (px)",
            "Red In Red Total Intensity 1",
        ]
        route_payloads = (
            (
                "dashboard_bulk_export",
                {
                    "uuids": [first_uuid, second_uuid],
                    "_columns": columns,
                    "_unit": "px",
                    "_puncta_source_contour_count": "exactly_2",
                },
            ),
            (
                "display_export_files",
                {
                    "visible_uuids": [first_uuid, second_uuid],
                    "uuids": [first_uuid, second_uuid],
                    "_columns": columns,
                    "_unit": "px",
                    "_puncta_source_contour_count": "exactly_2",
                },
            ),
        )

        for route_name, payload in route_payloads:
            for export_format in ("csv", "xlsx"):
                with self.subTest(route=route_name, export_format=export_format):
                    response = self.client.post(
                        reverse(route_name),
                        data=json.dumps({**payload, "_export": export_format}),
                        content_type="application/json",
                    )

                    self.assertEqual(response.status_code, 200)
                    rows = (
                        self._csv_rows(response)
                        if export_format == "csv"
                        else self._xlsx_rows(response)
                    )
                    self.assertEqual(rows[0], expected_headers)
                    self.assertEqual(
                        rows[1:],
                        [
                            [
                                "combined_red_filter_first",
                                "2" if export_format == "csv" else 2,
                                "1.000" if export_format == "csv" else 1,
                                "5.000" if export_format == "csv" else 5,
                            ]
                        ],
                    )

    def test_dashboard_combined_intensity_export_supports_total_max_without_average(self):
        first_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="combined_total_max_first",
        )
        second_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="combined_total_max_second",
        )
        self._add_cell_stat(first_uuid)
        self._add_cell_stat(second_uuid)
        columns = self._intensity_columns(statistics=("total", "max"))

        for export_format in ("csv", "xlsx"):
            with self.subTest(export_format=export_format):
                response = self.client.post(
                    reverse("dashboard_bulk_export"),
                    data=json.dumps(
                        {
                            "uuids": [first_uuid, second_uuid],
                            "_export": export_format,
                            "_columns": columns,
                            "_unit": "px",
                        }
                    ),
                    content_type="application/json",
                )
                self.assertEqual(response.status_code, 200)
                headers = (
                    self._csv_rows(response)[0]
                    if export_format == "csv"
                    else self._xlsx_headers(response)
                )
                self.assertEqual(len(headers), 26)
                self.assertIn("File Name", headers)
                self.assertIn("Red In Red Total Intensity 1", headers)
                self.assertIn("Green In Green Max Intensity 3", headers)
                self.assertFalse(any("Average Intensity" in header for header in headers))

    def test_display_combined_intensity_export_supports_average_only(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_combined_average_only",
        )
        other_visible_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_combined_average_not_selected",
        )
        self._add_cell_stat(saved_uuid)
        columns = self._intensity_columns(statistics=("average",))

        for export_format in ("csv", "xlsx"):
            with self.subTest(export_format=export_format):
                response = self.client.post(
                    reverse("display_export_files"),
                    data=json.dumps(
                        {
                            "visible_uuids": [saved_uuid, other_visible_uuid],
                            "uuids": [saved_uuid],
                            "_export": export_format,
                            "_columns": columns,
                            "_unit": "px",
                        }
                    ),
                    content_type="application/json",
                )
                self.assertEqual(response.status_code, 200)
                headers = (
                    self._csv_rows(response)[0]
                    if export_format == "csv"
                    else self._xlsx_headers(response)
                )
                self.assertEqual(len(headers), 14)
                self.assertIn("Red In Red Average Intensity 1", headers)
                self.assertIn("Green In Green Average Intensity 3", headers)
                self.assertFalse(any("Total Intensity" in header for header in headers))
                self.assertFalse(any("Max Intensity" in header for header in headers))

    def test_dashboard_combined_export_filename_scope_tracks_metric_selection(self):
        first_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="dashboard_combined_metric_scope_first",
        )
        second_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="dashboard_combined_metric_scope_second",
        )
        self._add_cell_stat(first_uuid)
        self._add_cell_stat(second_uuid)

        cases = [
            ("csv", [first_uuid, second_uuid], ["red_in_red_total_intensity_1"], "selected", 2),
            ("xlsx", [first_uuid, second_uuid], ["red_in_red_total_intensity_1"], "selected", 2),
            (
                "csv",
                [first_uuid],
                list(USER_SELECTABLE_TABLE_FIELDS),
                "all",
                1,
            ),
            (
                "xlsx",
                [first_uuid],
                list(USER_SELECTABLE_TABLE_FIELDS),
                "all",
                1,
            ),
        ]
        with patch(
            "core.services.export_filenames.timezone.now",
            return_value=self._fixed_export_time(),
        ):
            for export_format, uuids, columns, expected_scope, file_count in cases:
                with self.subTest(
                    export_format=export_format,
                    expected_scope=expected_scope,
                ):
                    response = self.client.post(
                        reverse("dashboard_bulk_export"),
                        data=json.dumps(
                            {
                                "uuids": uuids,
                                "_export": export_format,
                                "_columns": columns,
                                "_unit": "px",
                            }
                        ),
                        content_type="application/json",
                    )

                    self.assertEqual(response.status_code, 200)
                    self.assertExactExportFilename(
                        response,
                        scope=expected_scope,
                        file_count=file_count,
                        extension=export_format,
                    )

    def test_display_combined_csv_export_uses_visible_order_not_request_order(self):
        first_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_combined_first",
        )
        second_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_combined_second",
        )
        self._add_cell_stat(first_uuid)
        self._add_cell_stat(second_uuid)

        response = self.client.post(
            reverse("display_export_files"),
            data=json.dumps(
                {
                    "visible_uuids": [first_uuid, second_uuid],
                    "uuids": [second_uuid, first_uuid],
                    "_export": "csv",
                    "_columns": ["red_in_red_total_intensity_1"],
                    "_unit": "px",
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertExportFilename(
            response,
            scope="selected",
            file_count=2,
            extension="csv",
        )
        rows = self._csv_rows(response)
        self.assertEqual(rows[0], ["File Name", "Cell ID", "Red In Red Total Intensity 1"])
        self.assertEqual(
            [row[0] for row in rows[1:]],
            ["display_combined_first", "display_combined_second"],
        )

    def test_display_combined_xlsx_export_filters_headers(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_combined_xlsx",
        )
        other_visible_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_combined_xlsx_not_selected",
        )
        self._add_cell_stat(saved_uuid)

        response = self.client.post(
            reverse("display_export_files"),
            data=json.dumps(
                {
                    "visible_uuids": [saved_uuid, other_visible_uuid],
                    "uuids": [saved_uuid],
                    "_export": "xlsx",
                    "_columns": ["red_in_red_total_intensity_1"],
                    "_unit": "px",
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertExportFilename(
            response,
            scope="selected",
            file_count=1,
            extension="xlsx",
        )
        self.assertEqual(
            self._xlsx_headers(response),
            ["File Name", "Cell ID", "Red In Red Total Intensity 1"],
        )

    def test_display_combined_csv_export_respects_micron_unit_request(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_combined_um_export",
        )
        uploaded = UploadedImage.objects.get(uuid=saved_uuid)
        uploaded.scale_info = build_scale_info(
            manual_um_per_px=0.5, prefer_metadata=False
        )
        uploaded.save(update_fields=["scale_info"])
        self._add_cell_stat(saved_uuid)

        response = self.client.post(
            reverse("display_export_files"),
            data=json.dumps(
                {
                    "visible_uuids": [saved_uuid],
                    "uuids": [saved_uuid],
                    "_export": "csv",
                    "_columns": [
                        "puncta_distance",
                        "red_contour_1_center_xy",
                        "red_in_red_total_intensity_1",
                    ],
                    "_unit": "um",
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        rows = self._csv_rows(response)
        self.assertEqual(
            rows[0],
            [
                "File Name",
                "Cell ID",
                "Puncta Distance (µm)",
                "Red Contour 1 Center (x,y) (µm)",
                "Red In Red Total Intensity 1",
            ],
        )
        self.assertEqual(
            rows[1],
            ["display_combined_um_export", "1", "0.500", "5.000, 10.000", "5.000"],
        )

    def test_display_combined_export_filename_scope_tracks_metric_selection(self):
        first_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_combined_metric_scope_first",
        )
        second_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_combined_metric_scope_second",
        )
        self._add_cell_stat(first_uuid)
        self._add_cell_stat(second_uuid)

        cases = [
            ("csv", [first_uuid, second_uuid], ["red_in_red_total_intensity_1"], "selected", 2),
            ("xlsx", [first_uuid, second_uuid], ["red_in_red_total_intensity_1"], "selected", 2),
            (
                "csv",
                [first_uuid],
                list(USER_SELECTABLE_TABLE_FIELDS),
                "all",
                1,
            ),
            (
                "xlsx",
                [first_uuid],
                list(USER_SELECTABLE_TABLE_FIELDS),
                "all",
                1,
            ),
        ]
        with patch(
            "core.services.export_filenames.timezone.now",
            return_value=self._fixed_export_time(),
        ):
            for export_format, uuids, columns, expected_scope, file_count in cases:
                with self.subTest(
                    export_format=export_format,
                    expected_scope=expected_scope,
                ):
                    response = self.client.post(
                        reverse("display_export_files"),
                        data=json.dumps(
                            {
                                "visible_uuids": [first_uuid, second_uuid],
                                "uuids": uuids,
                                "_export": export_format,
                                "_columns": columns,
                                "_unit": "px",
                            }
                        ),
                        content_type="application/json",
                    )

                    self.assertEqual(response.status_code, 200)
                    self.assertExactExportFilename(
                        response,
                        scope=expected_scope,
                        file_count=file_count,
                        extension=export_format,
                    )

    def test_dashboard_combined_csv_export_respects_micron_unit_request(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="combined_um_export",
        )
        uploaded = UploadedImage.objects.get(uuid=saved_uuid)
        uploaded.scale_info = build_scale_info(
            manual_um_per_px=0.5, prefer_metadata=False
        )
        uploaded.save(update_fields=["scale_info"])
        self._add_cell_stat(saved_uuid)

        response = self.client.post(
            reverse("dashboard_bulk_export"),
            data=json.dumps(
                {
                    "uuids": [saved_uuid],
                    "_export": "csv",
                    "_columns": ["puncta_distance"],
                    "_unit": "um",
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        rows = self._csv_rows(response)
        self.assertEqual(rows[0], ["File Name", "Cell ID", "Puncta Distance (µm)"])
        self.assertEqual(rows[1], ["combined_um_export", "1", "0.500"])

    def test_combined_exports_reject_invalid_empty_and_inaccessible_requests(self):
        owned_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="combined_owned",
        )
        foreign_uuid = self._create_display_file(
            uploaded_owner=self.other_user,
            segmented_owner_id=self.other_user.id,
            filename="combined_foreign",
        )
        outside_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="combined_outside_visible",
        )
        self._add_cell_stat(owned_uuid)
        self._add_cell_stat(outside_uuid)

        no_files = self.client.post(
            reverse("dashboard_bulk_export"),
            data=json.dumps(
                {"uuids": [], "_export": "csv", "_columns": ["red_in_red_total_intensity_1"]}
            ),
            content_type="application/json",
        )
        invalid_columns = self.client.post(
            reverse("dashboard_bulk_export"),
            data=json.dumps({"uuids": [owned_uuid], "_export": "csv", "_columns": []}),
            content_type="application/json",
        )
        inaccessible = self.client.post(
            reverse("dashboard_bulk_export"),
            data=json.dumps(
                {
                    "uuids": [owned_uuid, foreign_uuid],
                    "_export": "csv",
                    "_columns": ["red_in_red_total_intensity_1"],
                }
            ),
            content_type="application/json",
        )
        not_visible = self.client.post(
            reverse("display_export_files"),
            data=json.dumps(
                {
                    "visible_uuids": [owned_uuid],
                    "uuids": [outside_uuid],
                    "_export": "csv",
                    "_columns": ["red_in_red_total_intensity_1"],
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(no_files.status_code, 400)
        self.assertEqual(invalid_columns.status_code, 400)
        self.assertEqual(inaccessible.status_code, 403)
        self.assertEqual(not_visible.status_code, 403)

    def test_combined_export_rejects_files_without_statistics(self):
        empty_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="combined_empty_stats",
        )

        response = self.client.post(
            reverse("dashboard_bulk_export"),
            data=json.dumps(
                {
                    "uuids": [empty_uuid],
                    "_export": "csv",
                    "_columns": ["red_in_red_total_intensity_1"],
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 400)

    def test_combined_export_excludes_deleted_cell_rows(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="combined_deleted_cells",
        )
        self._add_cell_stat(saved_uuid, cell_id=1)
        self._add_cell_stat(saved_uuid, cell_id=2)
        segmented = SegmentedImage.objects.get(UUID=saved_uuid)
        CellStatistics.objects.filter(segmented_image=segmented, cell_id=2).delete()

        response = self.client.post(
            reverse("dashboard_bulk_export"),
            data=json.dumps(
                {
                    "uuids": [saved_uuid],
                    "_export": "csv",
                    "_columns": ["red_in_red_total_intensity_1"],
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        rows = self._csv_rows(response)
        self.assertEqual([row[1] for row in rows[1:]], ["1"])

    def test_combined_nuclear_export_can_include_uncomputed_stats_as_na(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="combined_nuclear_filtered_export",
        )
        self._add_cell_stat(
            saved_uuid,
            properties={
                "selected_analysis": ["NuclearCellPairIntensity"],
                "nuclear_cell_pair_mode": "green_nucleus",
                "nuclear_cell_pair_status": "ok",
                "nuclear_cell_pair_contour_source": "canonical_slot_1",
            },
        )

        response = self.client.post(
            reverse("dashboard_bulk_export"),
            data=json.dumps(
                {
                    "uuids": [saved_uuid],
                    "_export": "csv",
                    "_columns": ["cell_pair_intensity_sum", "puncta_distance"],
                    "_unit": "px",
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        rows = self._csv_rows(response)
        self.assertEqual(
            rows[0],
            [
                "File Name",
                "Cell ID",
                "Puncta Distance (px)",
                "Measured Cell-Pair Intensity",
            ],
        )
        self.assertEqual(
            rows[1], ["combined_nuclear_filtered_export", "1", "N/A", "4.000"]
        )

    def test_display_csv_export_uses_generated_download_name(self):
        file_name = "display_csv_export_source"
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename=file_name,
        )
        self._add_cell_stat(saved_uuid)

        response = self.client.get(
            reverse("display", args=[saved_uuid]),
            {"_export": "csv"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertExportFilename(
            response,
            scope="all",
            file_count=1,
            extension="csv",
        )
        self.assertIn("text/csv", response["Content-Type"])

    def test_display_xlsx_export_uses_generated_download_name(self):
        file_name = "display_xlsx_export_source"
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename=file_name,
        )
        self._add_cell_stat(saved_uuid)

        response = self.client.get(
            reverse("display", args=[saved_uuid]),
            {"_export": "xlsx"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertExportFilename(
            response,
            scope="all",
            file_count=1,
            extension="xlsx",
        )
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )

    def test_display_single_export_filename_scope_tracks_metric_selection(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_filename_metric_scope",
        )
        other_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_filename_other_visible",
        )
        self._add_cell_stat(saved_uuid)
        self._add_cell_stat(other_uuid)

        cases = [
            ("csv", {}, "all", "csv"),
            ("xlsx", {}, "all", "xlsx"),
            (
                "csv",
                {"_columns": self._all_metric_columns()},
                "all",
                "csv",
            ),
            (
                "xlsx",
                {"_columns": self._all_metric_columns()},
                "all",
                "xlsx",
            ),
            (
                "csv",
                {"_columns": "red_in_red_total_intensity_1,puncta_distance"},
                "selected",
                "csv",
            ),
            (
                "xlsx",
                {"_columns": "red_in_red_total_intensity_1,puncta_distance"},
                "selected",
                "xlsx",
            ),
        ]
        with patch(
            "core.services.export_filenames.timezone.now",
            return_value=self._fixed_export_time(),
        ):
            for export_format, extra_params, expected_scope, extension in cases:
                with self.subTest(export_format=export_format, params=extra_params):
                    response = self.client.get(
                        reverse("display", args=[f"{saved_uuid},{other_uuid}"]),
                        {
                            "_export": export_format,
                            **extra_params,
                        },
                    )

                    self.assertEqual(response.status_code, 200)
                    self.assertExactExportFilename(
                        response,
                        scope=expected_scope,
                        file_count=1,
                        extension=extension,
                    )

    def test_display_csv_and_xlsx_exports_filter_selected_columns(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_filtered_export",
        )
        self._add_cell_stat(saved_uuid)

        csv_response = self.client.get(
            reverse("display", args=[saved_uuid]),
            {
                "_export": "csv",
                "_columns": (
                    "red_in_red_total_intensity_1,red_contour_1_center_xy,"
                    "measurement_contour_ratio_1"
                ),
            },
        )
        xlsx_response = self.client.get(
            reverse("display", args=[saved_uuid]),
            {
                "_export": "xlsx",
                "_columns": (
                    "red_in_red_total_intensity_1,red_contour_1_center_xy,"
                    "measurement_contour_ratio_1"
                ),
            },
        )

        self.assertEqual(csv_response.status_code, 200)
        self.assertEqual(xlsx_response.status_code, 200)
        expected_headers = [
            "Cell ID",
            "Red Contour 1 Center (x,y) (px)",
            "Red In Red Total Intensity 1",
            "Measurement/Contour Ratio 1 (Green/Red)",
        ]
        self.assertEqual(self._csv_rows(csv_response)[0], expected_headers)
        self.assertNotIn("Red In Red Max Intensity 1", self._csv_rows(csv_response)[0])
        self.assertNotIn("Red In Red Average Intensity 1", self._csv_rows(csv_response)[0])
        self.assertEqual(
            self._csv_rows(csv_response)[1],
            ["1", "10.000, 20.000", "5.000", "1.200"],
        )
        self.assertEqual(self._xlsx_headers(xlsx_response), expected_headers)
        self.assertNotIn("Red In Red Max Intensity 1", self._xlsx_headers(xlsx_response))
        self.assertNotIn("Red In Red Average Intensity 1", self._xlsx_headers(xlsx_response))

    def test_display_csv_and_xlsx_exports_respect_micron_unit_request(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_export_um",
        )
        uploaded = UploadedImage.objects.get(uuid=saved_uuid)
        uploaded.scale_info = build_scale_info(
            manual_um_per_px=0.5, prefer_metadata=False
        )
        uploaded.save(update_fields=["scale_info"])
        self._add_cell_stat(saved_uuid)

        selected_columns = (
            "puncta_distance,red_contour_1_center_xy,"
            "red_in_red_total_intensity_1,measurement_contour_ratio_1"
        )
        csv_response = self.client.get(
            reverse("display", args=[saved_uuid]),
            {
                "_export": "csv",
                "_unit": "um",
                "_columns": selected_columns,
            },
        )
        xlsx_response = self.client.get(
            reverse("display", args=[saved_uuid]),
            {
                "_export": "xlsx",
                "_unit": "um",
                "_columns": selected_columns,
            },
        )

        self.assertEqual(csv_response.status_code, 200)
        self.assertEqual(xlsx_response.status_code, 200)
        expected_headers = [
            "Cell ID",
            "Distance Between Red Puncta (µm)",
            "Red Contour 1 Center (x,y) (µm)",
            "Red In Red Total Intensity 1",
            "Measurement/Contour Ratio 1 (Green/Red)",
        ]
        self.assertEqual(self._csv_rows(csv_response)[0], expected_headers)
        self.assertEqual(
            self._csv_rows(csv_response)[1],
            ["1", "0.500", "5.000, 10.000", "5.000", "1.200"],
        )
        self.assertEqual(self._xlsx_headers(xlsx_response), expected_headers)
        self.assertEqual(
            self._xlsx_rows(xlsx_response)[1],
            [1, 0.5, "5.000, 10.000", 5, 1.2],
        )

    def test_filtered_exports_reject_invalid_or_empty_columns(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="display_invalid_columns",
        )
        self._add_cell_stat(saved_uuid)

        display_response = self.client.get(
            reverse("display", args=[saved_uuid]),
            {"_export": "csv", "_columns": "cell_id,unknown"},
        )
        dashboard_response = self.client.get(
            reverse("dashboard"),
            {"file_uuid": saved_uuid, "_export": "csv", "_columns": ""},
        )

        self.assertEqual(display_response.status_code, 400)
        self.assertEqual(dashboard_response.status_code, 400)

    def test_nuclear_export_can_include_uncomputed_stats_as_na(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="nuclear_filtered_export",
        )
        self._add_cell_stat(
            saved_uuid,
            properties={
                "selected_analysis": ["NuclearCellPairIntensity"],
                "nuclear_cell_pair_mode": "green_nucleus",
                "nuclear_cell_pair_status": "ok",
                "nuclear_cell_pair_contour_source": "canonical_slot_1",
            },
        )

        response = self.client.get(
            reverse("dashboard"),
            {
                "file_uuid": saved_uuid,
                "_export": "csv",
                "_columns": "cell_pair_intensity_sum,puncta_distance",
            },
        )

        self.assertEqual(response.status_code, 200)
        rows = self._csv_rows(response)
        self.assertEqual(
            rows[0],
            [
                "Cell ID",
                "Distance Between Red Puncta (px)",
                "Red Cell-Pair Intensity",
            ],
        )
        self.assertEqual(rows[1], ["1", "N/A", "4.000"])

    def test_green_red_intensity_disabled_export_does_not_emit_fake_values(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="green_red_disabled_filtered_export",
        )
        self._add_cell_stat(
            saved_uuid,
            properties={
                "selected_analysis": ["NuclearCellPairIntensity"],
                "nuclear_cell_pair_mode": "green_nucleus",
                "nuclear_cell_pair_status": "ok",
                "nuclear_cell_pair_contour_source": "canonical_slot_1",
            },
        )

        response = self.client.get(
            reverse("dashboard"),
            {
                "file_uuid": saved_uuid,
                "_export": "csv",
                "_columns": "red_in_red_total_intensity_1,red_in_red_max_intensity_1",
            },
        )

        self.assertEqual(response.status_code, 200)
        rows = self._csv_rows(response)
        self.assertEqual(
            rows[0],
            [
                "Cell ID",
                "Red In Red Total Intensity 1",
                "Red In Red Max Intensity 1",
            ],
        )
        self.assertEqual(rows[1], ["1", "N/A", "N/A"])

    def test_display_save_endpoint_is_idempotent_for_saved_file(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="already_saved",
        )

        response = self.client.post(
            reverse("display_save_files"),
            data=json.dumps({"uuids": [saved_uuid]}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["saved_count"], 0)
        self.assertEqual(payload["already_saved_count"], 1)
        self.assertEqual(payload["already_saved_uuids"], [saved_uuid])
        self.assertEqual(
            SegmentedImage.objects.get(UUID=saved_uuid).user_id,
            self.user.id,
        )

    def test_display_save_endpoint_handles_mixed_saved_and_transient_selection(self):
        transient_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.guest_user_id,
            filename="mixed_transient",
        )
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="mixed_saved",
        )
        self._set_transient_uuids([transient_uuid])

        response = self.client.post(
            reverse("display_save_files"),
            data=json.dumps({"uuids": [saved_uuid, transient_uuid]}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["saved_count"], 1)
        self.assertEqual(payload["already_saved_count"], 1)
        self.assertIn(transient_uuid, payload["saved_uuids"])
        self.assertIn(saved_uuid, payload["already_saved_uuids"])
        self.assertEqual(
            SegmentedImage.objects.get(UUID=transient_uuid).user_id,
            self.user.id,
        )
        self.assertEqual(
            SegmentedImage.objects.get(UUID=saved_uuid).user_id,
            self.user.id,
        )

        session = self.client.session
        self.assertNotIn(transient_uuid, session.get("transient_experiment_uuids", []))

    def test_display_save_endpoint_rejects_when_storage_quota_is_exceeded(self):
        transient_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.guest_user_id,
            filename="quota_transient",
        )
        self._set_transient_uuids([transient_uuid])
        self.user.total_storage = 32
        self.user.save(update_fields=["total_storage"])

        with TemporaryDirectory() as temp_media:
            with override_settings(MEDIA_ROOT=temp_media):
                self._write_run_bytes(temp_media, transient_uuid, size=96)
                response = self.client.post(
                    reverse("display_save_files"),
                    data=json.dumps({"uuids": [transient_uuid]}),
                    content_type="application/json",
                )

        self.assertEqual(response.status_code, 507)
        payload = response.json()
        self.assertEqual(payload["code"], "storage_full")
        self.assertEqual(
            payload["error"],
            "Selected files could not be saved because your storage is full. Free up space and try again.",
        )
        self.assertEqual(
            SegmentedImage.objects.get(UUID=transient_uuid).user_id,
            self.guest_user_id,
        )
        self.assertIn(
            transient_uuid,
            self.client.session.get("transient_experiment_uuids", []),
        )

    def test_display_unsave_endpoint_rejects_invalid_payload(self):
        response = self.client.post(
            reverse("display_unsave_files"),
            data=json.dumps({"uuids": "bad-shape"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 400)

    def test_display_unsave_endpoint_unsaves_saved_file_and_adds_transient(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="manual_unsave_candidate",
        )

        response = self.client.post(
            reverse("display_unsave_files"),
            data=json.dumps({"uuids": [saved_uuid]}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["unsaved_count"], 1)
        self.assertEqual(payload["already_unsaved_count"], 0)
        self.assertEqual(payload["unsaved_uuids"], [saved_uuid])
        self.assertEqual(
            SegmentedImage.objects.get(UUID=saved_uuid).user_id,
            self.guest_user_id,
        )
        session = self.client.session
        self.assertIn(saved_uuid, session.get("transient_experiment_uuids", []))

        dashboard_response = self.client.get(reverse("dashboard"))
        self.assertEqual(dashboard_response.status_code, 200)
        self.assertNotContains(dashboard_response, "manual_unsave_candidate")

    def test_display_unsave_endpoint_rejects_foreign_or_unavailable_uuid(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="owner_saved",
        )
        foreign_uuid = self._create_display_file(
            uploaded_owner=self.other_user,
            segmented_owner_id=self.other_user.id,
            filename="foreign_saved",
        )

        response = self.client.post(
            reverse("display_unsave_files"),
            data=json.dumps({"uuids": [saved_uuid, foreign_uuid]}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(
            SegmentedImage.objects.get(UUID=saved_uuid).user_id, self.user.id
        )

    def test_display_unsave_endpoint_is_idempotent_for_already_unsaved_file(self):
        transient_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.guest_user_id,
            filename="already_unsaved",
        )
        self._set_transient_uuids([transient_uuid])

        response = self.client.post(
            reverse("display_unsave_files"),
            data=json.dumps({"uuids": [transient_uuid]}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["unsaved_count"], 0)
        self.assertEqual(payload["already_unsaved_count"], 1)
        self.assertEqual(payload["already_unsaved_uuids"], [transient_uuid])
        self.assertEqual(
            SegmentedImage.objects.get(UUID=transient_uuid).user_id,
            self.guest_user_id,
        )

    def test_display_unsave_endpoint_handles_mixed_saved_and_unsaved_selection(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="mixed_saved_unsave",
        )
        transient_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.guest_user_id,
            filename="mixed_unsaved_unsave",
        )
        self._set_transient_uuids([transient_uuid])

        response = self.client.post(
            reverse("display_unsave_files"),
            data=json.dumps({"uuids": [saved_uuid, transient_uuid]}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["unsaved_count"], 1)
        self.assertEqual(payload["already_unsaved_count"], 1)
        self.assertIn(saved_uuid, payload["unsaved_uuids"])
        self.assertIn(transient_uuid, payload["already_unsaved_uuids"])
        self.assertEqual(
            SegmentedImage.objects.get(UUID=saved_uuid).user_id,
            self.guest_user_id,
        )

    def test_display_sync_selection_rejects_selected_not_in_visible_list(self):
        visible_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.guest_user_id,
            filename="visible_sync",
        )
        outside_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.guest_user_id,
            filename="outside_sync",
        )
        self._set_transient_uuids([visible_uuid, outside_uuid])

        response = self.client.post(
            reverse("display_sync_file_selection"),
            data=json.dumps(
                {
                    "visible_uuids": [visible_uuid],
                    "selected_uuids": [outside_uuid],
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 400)

    def test_display_sync_selection_applies_save_and_unsave_together(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="sync_saved",
        )
        transient_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.guest_user_id,
            filename="sync_transient",
        )
        self._set_transient_uuids([transient_uuid])

        response = self.client.post(
            reverse("display_sync_file_selection"),
            data=json.dumps(
                {
                    "visible_uuids": [saved_uuid, transient_uuid],
                    "selected_uuids": [transient_uuid],
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["saved_count"], 1)
        self.assertEqual(payload["unsaved_count"], 1)
        self.assertIn(transient_uuid, payload["saved_uuids"])
        self.assertIn(saved_uuid, payload["unsaved_uuids"])
        self.assertEqual(
            SegmentedImage.objects.get(UUID=transient_uuid).user_id,
            self.user.id,
        )
        self.assertEqual(
            SegmentedImage.objects.get(UUID=saved_uuid).user_id,
            self.guest_user_id,
        )

        session = self.client.session
        transient_session = set(session.get("transient_experiment_uuids", []))
        self.assertIn(saved_uuid, transient_session)
        self.assertNotIn(transient_uuid, transient_session)

        dashboard_response = self.client.get(reverse("dashboard"))
        self.assertEqual(dashboard_response.status_code, 200)
        self.assertContains(dashboard_response, "sync_transient")
        self.assertNotContains(dashboard_response, "sync_saved")

    def test_display_sync_selection_allows_net_save_when_unsave_frees_space(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="sync_saved_quota",
        )
        transient_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.guest_user_id,
            filename="sync_transient_quota",
        )
        self._set_transient_uuids([transient_uuid])
        self.user.total_storage = 80
        self.user.save(update_fields=["total_storage"])

        with TemporaryDirectory() as temp_media:
            with override_settings(MEDIA_ROOT=temp_media):
                self._write_run_bytes(temp_media, saved_uuid, size=80)
                self._write_run_bytes(temp_media, transient_uuid, size=60)
                response = self.client.post(
                    reverse("display_sync_file_selection"),
                    data=json.dumps(
                        {
                            "visible_uuids": [saved_uuid, transient_uuid],
                            "selected_uuids": [transient_uuid],
                        }
                    ),
                    content_type="application/json",
                )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["saved_count"], 1)
        self.assertEqual(payload["unsaved_count"], 1)
        self.assertEqual(
            SegmentedImage.objects.get(UUID=transient_uuid).user_id,
            self.user.id,
        )
        self.assertEqual(
            SegmentedImage.objects.get(UUID=saved_uuid).user_id,
            self.guest_user_id,
        )

    def test_display_sync_selection_rejects_atomically_when_net_quota_is_exceeded(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="sync_saved_atomic",
        )
        transient_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.guest_user_id,
            filename="sync_transient_atomic",
        )
        self._set_transient_uuids([transient_uuid])
        self.user.total_storage = 30
        self.user.save(update_fields=["total_storage"])

        with TemporaryDirectory() as temp_media:
            with override_settings(MEDIA_ROOT=temp_media):
                self._write_run_bytes(temp_media, saved_uuid, size=20)
                self._write_run_bytes(temp_media, transient_uuid, size=60)
                response = self.client.post(
                    reverse("display_sync_file_selection"),
                    data=json.dumps(
                        {
                            "visible_uuids": [saved_uuid, transient_uuid],
                            "selected_uuids": [transient_uuid],
                        }
                    ),
                    content_type="application/json",
                )

        self.assertEqual(response.status_code, 507)
        payload = response.json()
        self.assertEqual(payload["code"], "storage_full")
        self.assertEqual(
            SegmentedImage.objects.get(UUID=transient_uuid).user_id,
            self.guest_user_id,
        )
        self.assertEqual(
            SegmentedImage.objects.get(UUID=saved_uuid).user_id,
            self.user.id,
        )
        transient_session = set(
            self.client.session.get("transient_experiment_uuids", [])
        )
        self.assertIn(transient_uuid, transient_session)
        self.assertNotIn(saved_uuid, transient_session)

    def test_display_sync_selection_rejects_foreign_visible_file(self):
        owned_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.guest_user_id,
            filename="sync_owned",
        )
        foreign_uuid = self._create_display_file(
            uploaded_owner=self.other_user,
            segmented_owner_id=self.guest_user_id,
            filename="sync_foreign",
        )
        self._set_transient_uuids([owned_uuid, foreign_uuid])

        response = self.client.post(
            reverse("display_sync_file_selection"),
            data=json.dumps(
                {
                    "visible_uuids": [owned_uuid, foreign_uuid],
                    "selected_uuids": [owned_uuid],
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 403)

    def test_display_view_respects_channel_visibility_preference(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="visibility_saved",
        )
        prefs = get_user_preferences(self.user)
        prefs["show_saved_file_channels"] = False
        update_user_preferences(self.user, prefs)

        response = self.client.get(reverse("display", args=[saved_uuid]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="sidebar channels-hidden"')
        self.assertContains(response, "Show Channels")

    def test_preprocess_view_respects_channel_visibility_preference(self):
        preprocess_uuid = self._create_preprocess_file(filename="visibility_preprocess")
        prefs = get_user_preferences(self.user)
        prefs["show_saved_file_channels"] = False
        update_user_preferences(self.user, prefs)

        response = self.client.get(reverse("pre_process", args=[preprocess_uuid]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="sidebar channels-hidden"')
        self.assertContains(response, "Show Channels")

    def test_display_view_respects_scale_visibility_preference(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="scale_visibility_saved",
        )
        prefs = get_user_preferences(self.user)
        prefs["show_saved_file_scales"] = False
        update_user_preferences(self.user, prefs)

        response = self.client.get(reverse("display", args=[saved_uuid]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="sidebar scales-hidden"')
        self.assertContains(response, "Show Scale")

    def test_preprocess_view_respects_scale_visibility_preference(self):
        preprocess_uuid = self._create_preprocess_file(
            filename="scale_visibility_preprocess"
        )
        prefs = get_user_preferences(self.user)
        prefs["show_saved_file_scales"] = False
        update_user_preferences(self.user, prefs)

        response = self.client.get(reverse("pre_process", args=[preprocess_uuid]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="sidebar scales-hidden"')
        self.assertContains(response, "Show Scale")

    def test_display_dashboard_and_preprocess_use_sidebar_spatial_unit_preference(self):
        saved_uuid = self._create_display_file(
            uploaded_owner=self.user,
            segmented_owner_id=self.user.id,
            filename="sidebar_unit_saved",
        )
        preprocess_uuid = self._create_preprocess_file(
            filename="sidebar_unit_preprocess"
        )
        prefs = get_user_preferences(self.user)
        prefs["sidebar_spatial_stats_unit"] = "um"
        update_user_preferences(self.user, prefs)

        display_response = self.client.get(reverse("display", args=[saved_uuid]))
        dashboard_response = self.client.get(reverse("dashboard"))
        preprocess_response = self.client.get(
            reverse("pre_process", args=[preprocess_uuid])
        )

        self.assertContains(
            display_response, '"initialSidebarSpatialStatsUnit": "um"', html=False
        )
        self.assertContains(
            dashboard_response,
            '"initialSidebarSpatialStatsUnit": "um"',
            html=False,
        )
        self.assertContains(
            preprocess_response,
            '"initialSidebarSpatialStatsUnit": "um"',
            html=False,
        )

    def test_preprocess_post_rejects_tampered_scale_uuid_map(self):
        preprocess_uuid = self._create_preprocess_file(filename="tamper_preprocess")
        outside_uuid = self._create_preprocess_file(filename="outside_preprocess")

        response = self.client.post(
            reverse("pre_process", args=[preprocess_uuid]),
            data={"file_scale_map": json.dumps({outside_uuid: 0.2})},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 403)

    def test_preprocess_post_rejects_tampered_scale_revert_uuid_map(self):
        preprocess_uuid = self._create_preprocess_file(
            filename="tamper_revert_preprocess"
        )
        outside_uuid = self._create_preprocess_file(
            filename="outside_revert_preprocess"
        )

        response = self.client.post(
            reverse("pre_process", args=[preprocess_uuid]),
            data={"file_scale_revert_uuids": json.dumps([outside_uuid])},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 403)

    @patch(
        "core.views.pre_process.run_analysis_batch",
        return_value=SimpleNamespace(storage_warning_message=""),
    )
    def test_preprocess_post_persists_manual_scale_override_before_analysis(
        self,
        mock_run_analysis_batch,
    ):
        preprocess_uuid = self._create_preprocess_file(
            filename="scale_override_preprocess"
        )

        response = self.client.post(
            reverse("pre_process", args=[preprocess_uuid]),
            data={"file_scale_map": json.dumps({preprocess_uuid: 0.27})},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.json()["redirect"], reverse("display", args=[preprocess_uuid])
        )
        mock_run_analysis_batch.assert_called_once()
        uploaded = UploadedImage.objects.get(uuid=preprocess_uuid)
        scale_info = uploaded.scale_info or {}
        self.assertEqual(scale_info.get("source"), "manual_override")
        self.assertAlmostEqual(
            float(scale_info.get("effective_um_per_px", 0)), 0.27, places=6
        )

    @patch(
        "core.views.pre_process.run_analysis_batch",
        return_value=SimpleNamespace(storage_warning_message=""),
    )
    def test_preprocess_post_reverts_manual_override_to_metadata_scale(
        self,
        mock_run_analysis_batch,
    ):
        preprocess_uuid = self._create_preprocess_file(
            filename="scale_revert_preprocess"
        )
        uploaded = UploadedImage.objects.get(uuid=preprocess_uuid)
        uploaded.scale_info = apply_manual_override_scale(
            build_scale_info(
                manual_um_per_px=0.2,
                prefer_metadata=True,
                metadata_um_per_px=0.11,
                status="ok",
            ),
            effective_um_per_px=0.27,
        )
        uploaded.save(update_fields=["scale_info"])

        response = self.client.post(
            reverse("pre_process", args=[preprocess_uuid]),
            data={"file_scale_revert_uuids": json.dumps([preprocess_uuid])},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.json()["redirect"], reverse("display", args=[preprocess_uuid])
        )
        mock_run_analysis_batch.assert_called_once()
        uploaded.refresh_from_db()
        scale_info = uploaded.scale_info or {}
        self.assertEqual(scale_info.get("source"), "metadata")
        self.assertAlmostEqual(
            float(scale_info.get("effective_um_per_px", 0)), 0.11, places=6
        )


class ChannelVisibilityPreferenceTests(TestCase):
    def setUp(self):
        self.client = Client()
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            email="visibility@example.com",
            password="TestPass123!",
        )
        self.assertTrue(
            self.client.login(
                email="visibility@example.com",
                password="TestPass123!",
            )
        )

    def _create_saved_sidebar_file(self, filename: str = "sidebar_saved") -> str:
        file_uuid = uuid4()
        UploadedImage.objects.create(
            user=self.user,
            name=filename,
            uuid=file_uuid,
            file_location=f"{file_uuid}/{filename}.dv",
        )
        SegmentedImage.objects.create(
            user=self.user,
            UUID=file_uuid,
            file_location=f"user_{file_uuid}/{filename}.png",
            ImagePath=f"{file_uuid}/output/{filename}_frame_0.png",
            CellPairPrefix=f"{file_uuid}/segmented/cell_",
            NumCells=1,
        )
        return str(file_uuid)

    def _create_preprocess_sidebar_file(
        self, filename: str = "sidebar_preprocess"
    ) -> str:
        file_uuid = uuid4()
        uploaded = UploadedImage.objects.create(
            user=self.user,
            name=filename,
            uuid=file_uuid,
            file_location=f"{file_uuid}/{filename}.dv",
        )
        DVLayerTifPreview.objects.create(
            uploaded_image_uuid=uploaded,
            wavelength="DAPI",
            file_location=f"{file_uuid}/{filename}_preview.png",
        )
        return str(file_uuid)

    def _build_experiment_workflow_defaults_payload(self) -> dict[str, object]:
        return {
            "selected_plugins": ["PunctaDistance", "GreenRedIntensity"],
            "module_enabled": True,
            "enforce_layer_count": True,
            "enforce_wavelengths": False,
            "show_legacy_plugins": False,
            "manual_required_channels": ["channel_blue"],
            "green_contour_filter_enabled": True,
            "green_dot_split_enabled": False,
            "green_dot_split_mode": "aggressive",
            "alternate_red_detection": True,
            "puncta_line_width": 2.5,
            "puncta_line_width_unit": "um",
            "cen_dot_distance": 11.2,
            "cen_dot_distance_unit": "px",
            "cen_dot_proximity_radius": 6.5,
            "cen_dot_proximity_radius_unit": "um",
            "biorientation_red_min_distance": 1.5,
            "biorientation_red_min_distance_unit": "px",
            "biorientation_red_max_distance": 44.5,
            "biorientation_red_max_distance_unit": "um",
            "biorientation_collinearity_threshold": 77,
            "puncta_line_mode": "green_puncta",
            "nuclear_cell_pair_mode": "red_nucleus",
            "nuclear_cell_pair_contour_mode": "aggressive",
            "microns_per_pixel": 0.25,
            "use_metadata_scale": False,
            "use_metadata_channel_order": False,
            "fallback_channel_order": [
                "channel_green",
                "DIC",
                "channel_red",
                "channel_blue",
            ],
        }

    def test_preferences_page_renders_review_modal_and_form_review_hooks(self):
        response = self.client.get(reverse("workflow_defaults"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="workflowDefaultsNav"', html=False)
        self.assertContains(response, "Workflow Defaults")
        self.assertContains(response, "css/components/workflow-controls.css", html=False)
        self.assertContains(response, "css/pages/workflow-defaults.css", html=False)
        content = response.content.decode("utf-8")
        self.assertLess(
            content.index("css/components/workflow-controls.css"),
            content.index("css/pages/workflow-defaults.css"),
        )
        self.assertContains(
            response, 'id="pluginForm" data-review-section="plugins"', html=False
        )
        self.assertContains(
            response, 'id="advancedForm" data-review-section="advanced"', html=False
        )
        self.assertContains(
            response, 'id="savingForm" data-review-section="saving"', html=False
        )
        self.assertContains(
            response, 'data-workflow-card="plugin-required-channels"', html=False
        )
        self.assertContains(
            response, 'data-workflow-card="validation-enforcement"', html=False
        )
        self.assertContains(
            response, 'data-workflow-card="saving-preferences"', html=False
        )
        self.assertContains(
            response, 'data-workflow-card="sidebar-preferences"', html=False
        )
        self.assertContains(response, 'data-workflow-action-card="plugins"', html=False)
        self.assertContains(
            response, 'data-workflow-action-card="advanced"', html=False
        )
        self.assertContains(response, 'data-workflow-action-card="saving"', html=False)
        self.assertContains(response, 'id="advancedOptionalChecksNote"', html=False)
        self.assertContains(response, 'id="advancedOptionalChecksGroup"', html=False)
        self.assertContains(response, 'id="advancedManualChannelsGroup"', html=False)
        self.assertContains(response, 'id="manualRequiredChannels"', html=False)
        self.assertContains(response, 'id="advancedLayerCheckRow"', html=False)
        self.assertContains(response, 'id="advancedWavelengthCheckRow"', html=False)
        self.assertContains(response, 'id="sidebar_starts_open"', html=False)
        self.assertContains(response, 'id="prefsGfpFilterExperimentalDot"', html=False)
        self.assertContains(response, 'id="dot_split_enabled"', html=False)
        self.assertContains(response, 'id="dot_split_target"', html=False)
        self.assertContains(response, 'id="green_dot_split_enabled"', html=False)
        self.assertContains(response, 'id="green_dot_split_mode"', html=False)
        self.assertContains(response, 'id="red_dot_split_enabled"', html=False)
        self.assertContains(response, 'id="red_dot_split_mode"', html=False)
        self.assertContains(response, "Wavelength Channel Assignment")
        self.assertContains(response, 'id="use_metadata_channel_order"', html=False)
        self.assertContains(response, 'id="prefsFallbackChannelOrder"', html=False)
        self.assertContains(
            response, '<div class="channel-plane-row" aria-hidden="true">', html=False
        )
        self.assertContains(
            response,
            '<div class="channel-bar" id="prefsFallbackChannelOrder"',
            html=False,
        )
        self.assertContains(response, '<span class="channel-chip', html=False)
        self.assertContains(response, "sortablejs@1.15.0/Sortable.min.js")
        self.assertNotContains(response, "channel-order-chip")
        self.assertNotContains(response, "channel-order-bar")
        self.assertContains(response, "Split Merged Dots")
        self.assertNotContains(response, "Split Merged Green Dots")
        self.assertNotContains(response, "Split Merged Red Dots")
        self.assertContains(response, "Signal Quantification")
        self.assertContains(response, 'id="signal_quantification_enabled"', html=False)
        self.assertContains(response, 'id="signal_quantification_mode"', html=False)
        self.assertContains(response, "Red/Green Contour Intensities")
        self.assertContains(response, "Alternate Nucleus Detection")
        self.assertContains(response, "Nucleus Contour Mode")
        self.assertContains(response, 'id="nuclear_cell_pair_contour_mode"', html=False)
        self.assertContains(response, 'id="nuclearContourModeRow"', html=False)
        self.assertContains(response, "nuclear-contour-mode-child")
        self.assertContains(
            response, 'id="nuclear_cell_pair_contour_mode_value"', html=False
        )
        self.assertContains(
            response,
            "Controls the primary Red/Green signal measurement workflow for this experiment. Choose one primary mode: Puncta Distance or Nuclear, Cell-Pair Intensity.",
        )
        self.assertContains(
            response,
            "Primary Mode selects which mutually exclusive signal workflow is saved as the default.",
        )
        self.assertContains(
            response,
            "Detects the first two usable puncta in the selected source channel, measures the distance between their centers",
        )
        self.assertContains(
            response,
            "Measures signal from the selected measurement channel inside the selected nucleus contour and inside the full DIC cell-pair mask.",
        )
        self.assertContains(
            response,
            "Red Nucleus uses alternate Red detection, and Green Nucleus uses alternate Green detection.",
        )
        self.assertContains(
            response,
            "All other stat modules enabled in Puncta Distance mode.",
        )
        self.assertContains(
            response, 'id="alternate_nucleus_detection_enabled"', html=False
        )
        self.assertNotContains(response, 'id="cell_parentage_mode"', html=False)
        self.assertNotContains(response, "Mother/Daughter Mode")
        self.assertNotContains(
            response, 'id="biorientation_green_split_enabled"', html=False
        )
        self.assertNotContains(response, 'id="alternate_red_detection"', html=False)
        self.assertNotContains(
            response, 'data-workflow-card="channel-requirements"', html=False
        )
        html = response.content.decode()
        plugin_defaults_index = html.index(
            '<div class="card" data-workflow-card="plugin-defaults"'
        )
        dot_detection_index = html.index(
            '<div class="card" data-workflow-card="dot-detection"'
        )
        measurement_scale_index = html.index(
            '<div class="card" data-workflow-card="measurement-scale"'
        )
        self.assertLess(plugin_defaults_index, dot_detection_index)
        self.assertLess(dot_detection_index, measurement_scale_index)
        alternate_detection_index = html.index(
            'id="alternate_nucleus_detection_enabled"'
        )
        contour_mode_row_index = html.index('id="nuclearContourModeRow"')
        self.assertLess(alternate_detection_index, contour_mode_row_index)
        advanced_plugin_behavior = html[
            html.index(
                '<div class="card" data-workflow-card="advanced-plugin-behavior"'
            ) : html.index(
                '<div class="card" data-workflow-card="validation-enforcement"'
            )
        ]
        self.assertIn("Show Legacy Blue-Channel Plugins", advanced_plugin_behavior)
        self.assertNotIn("Filter Green Contours", advanced_plugin_behavior)
        self.assertNotIn("Split Merged Dots", advanced_plugin_behavior)
        self.assertNotIn("Enable Alternate Red Detection", advanced_plugin_behavior)
        self.assertContains(
            response,
            "Start Sidebars Open On Dashboard, Display, And Preprocess",
        )
        self.assertContains(
            response,
            "Stats-required channels always stay enforced. Manual channel requirements and all-channel enforcement are saved here and pause when this module is OFF.",
        )
        self.assertContains(
            response,
            "Require exactly four layers before preprocessing.",
        )
        self.assertContains(
            response,
            "Require DIC, Blue, Red, and Green even if not needed by selected statistics.",
        )
        self.assertContains(
            response,
            "Filters low-confidence Green signal contours in challenging images.",
        )
        self.assertContains(
            response,
            "If disabled, the standard nucleus contour path is used.",
        )
        self.assertContains(
            response,
            "Aggressive uses tighter speckle-derived alternate nucleus masks",
        )
        self.assertContains(response, 'id="reviewChangesBackdrop"', html=False)
        self.assertContains(
            response, 'class="review-backdrop popup-backdrop"', html=False
        )
        self.assertContains(response, 'class="review-modal popup-surface"', html=False)
        self.assertContains(response, 'id="reviewKeepOld"', html=False)
        self.assertContains(response, 'id="reviewConfirmChanges"', html=False)
        self.assertContains(response, 'id="leaveUnsavedBackdrop"', html=False)
        self.assertContains(response, 'id="leaveUnsavedKeepOld"', html=False)
        self.assertContains(response, 'id="leaveUnsavedConfirmNew"', html=False)
        self.assertContains(response, 'id="leaveUnsavedListWrap"', html=False)
        self.assertContains(response, 'id="leaveUnsavedList"', html=False)
        self.assertContains(response, "Leave without saving changes?")
        self.assertContains(response, "Keep Old")
        self.assertContains(response, "Confirm Changes")
        self.assertContains(response, "Confirm New")

    def test_experiment_page_renders_workflow_default_save_controls(self):
        response = self.client.get(reverse("experiment"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="saveWorkflowDefaultsPrimary"', html=False)
        self.assertContains(response, 'id="saveWorkflowDefaultsAdvanced"', html=False)
        self.assertContains(response, 'id="saveWorkflowDefaultsBackdrop"', html=False)
        self.assertContains(response, 'id="saveWorkflowDefaultsConfirm"', html=False)
        self.assertContains(response, "Save as Workflow Default")
        self.assertContains(response, "Save as workflow default?")
        self.assertContains(response, "Keep Old Settings")
        self.assertContains(response, "Keep New Changes")
        self.assertNotContains(response, 'id="cellParentageModeInline"', html=False)
        self.assertNotContains(response, 'id="cellParentageModeMount"', html=False)
        self.assertContains(response, "sortablejs@1.15.0/Sortable.min.js")
        self.assertContains(response, "css/components/workflow-controls.css", html=False)
        self.assertContains(response, "css/pages/experiment.css", html=False)
        self.assertContains(response, "js/pages/experiment.js", html=False)
        content = response.content.decode("utf-8")
        self.assertLess(
            content.index("css/components/workflow-controls.css"),
            content.index("css/pages/experiment.css"),
        )
        experiment_source = _frontend_static_text("js/pages/experiment.js")
        self.assertIn(
            "fallbackChannelOrderBar.className = 'channel-bar';",
            experiment_source,
        )
        self.assertIn("planeRow.className = 'channel-plane-row';", experiment_source)
        self.assertIn("chip.className = `channel-chip", experiment_source)
        self.assertIn("window.Sortable.create(bar,", experiment_source)
        self.assertNotContains(response, "channel-order-chip")
        self.assertNotContains(response, "channel-order-bar")
        self.assertNotIn("channel-order-chip", experiment_source)
        self.assertNotIn("channel-order-bar", experiment_source)

    def test_experiment_page_contains_mode_aware_signal_quantification_info_text(self):
        response = self.client.get(reverse("experiment"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "js/pages/experiment.js", html=False)
        experiment_source = _frontend_static_text("js/pages/experiment.js")
        self.assertIn(
            "Signal Quantification controls the primary Red/Green signal measurement workflow for this experiment.",
            experiment_source,
        )
        self.assertIn(
            "Puncta Distance detects the first two usable puncta in the selected source channel",
            experiment_source,
        )
        self.assertIn(
            "Red/Green Contour Intensities optionally calculates raw intensity sums inside detected Red and Green contour masks.",
            experiment_source,
        )
        self.assertIn(
            "Nuclear, Cell-Pair Intensity measures signal from the selected measurement channel inside the selected nucleus contour",
            experiment_source,
        )
        self.assertIn(
            "Nucleus Contour Mode: Balanced keeps the current alternate nucleus contour behavior",
            experiment_source,
        )
        self.assertIn("nuclear-contour-mode-child", experiment_source)
        self.assertIn(
            "nuclearContourModeSelect.setDisabled(!statsState.alternateNucleusDetectionEnabled)",
            experiment_source,
        )
        self.assertNotIn(
            "Only affects Nuclear, Cell-Pair Intensity and uses the alternate contour detection path on the selected nucleus source channel only.",
            experiment_source,
        )
        self.assertNotIn(
            "Alternate Nucleus Detection: only affects Nuclear, Cell-Pair Intensity",
            experiment_source,
        )
        self.assertNotIn(
            "Alternate Nucleus Detection changes only the Nuclear, Cell-Pair nucleus contour path.",
            experiment_source,
        )
        alternate_detection_js_index = experiment_source.index(
            "alternateLabel.textContent = 'Alternate Nucleus Detection';"
        )
        contour_mode_js_index = experiment_source.index(
            "contourModeLabel.textContent = 'Nucleus Contour Mode:'"
        )
        self.assertLess(alternate_detection_js_index, contour_mode_js_index)
        self.assertIn("Required channels: Red and Green.", experiment_source)
        self.assertIn(
            "All other stat modules enabled in Puncta Distance mode.",
            experiment_source,
        )
        self.assertIn(
            "Nuclear, Cell-Pair Intensity primary mode on. Other stat modules disabled.",
            experiment_source,
        )
        self.assertIn("signalQuantificationInfoDot", experiment_source)
        self.assertIn("buildSignalQuantificationInfoText", experiment_source)
        self.assertContains(response, 'id="dotSplitEnabled"', html=False)
        self.assertContains(response, 'id="dotSplitTargetMount"', html=False)
        self.assertContains(response, 'id="greenDotSplitModeMount"', html=False)
        self.assertContains(response, 'id="redDotSplitModeMount"', html=False)
        self.assertContains(response, "Split Merged Dots")
        self.assertNotContains(response, "Split Merged Green Dots")
        self.assertNotContains(response, "Split Merged Red Dots")

    def test_experiment_workflow_defaults_endpoint_persists_popup_settings(self):
        payload = self._build_experiment_workflow_defaults_payload()

        response = self.client.post(
            reverse("experiment_workflow_defaults"),
            data=json.dumps(payload),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.json()["message"],
            "Workflow default saved. Future experiments will start with these settings.",
        )

        self.user.refresh_from_db()
        defaults = get_user_preferences(self.user)["experiment_defaults"]
        self.assertEqual(
            defaults["selected_plugins"], ["PunctaDistance", "GreenRedIntensity"]
        )
        self.assertTrue(defaults["module_enabled"])
        self.assertTrue(defaults["enforce_layer_count"])
        self.assertFalse(defaults["enforce_wavelengths"])
        self.assertFalse(defaults["show_legacy_plugins"])
        self.assertEqual(defaults["manual_required_channels"], ["channel_blue"])
        self.assertTrue(defaults["green_contour_filter_enabled"])
        self.assertFalse(defaults["green_dot_split_enabled"])
        self.assertEqual(defaults["green_dot_split_mode"], "aggressive")
        self.assertTrue(defaults["alternate_nucleus_detection_enabled"])
        self.assertTrue(defaults["alternate_red_detection"])
        self.assertEqual(defaults["puncta_line_width"], 2.5)
        self.assertEqual(defaults["puncta_line_width_unit"], "um")
        self.assertEqual(defaults["cen_dot_distance"], 11.2)
        self.assertEqual(defaults["cen_dot_distance_unit"], "px")
        self.assertEqual(defaults["cen_dot_proximity_radius"], 6.5)
        self.assertEqual(defaults["cen_dot_proximity_radius_unit"], "um")
        self.assertEqual(defaults["biorientation_red_min_distance"], 1.5)
        self.assertEqual(defaults["biorientation_red_min_distance_unit"], "px")
        self.assertEqual(defaults["biorientation_red_max_distance"], 44.5)
        self.assertEqual(defaults["biorientation_red_max_distance_unit"], "um")
        self.assertEqual(defaults["biorientation_collinearity_threshold"], 77)
        self.assertEqual(defaults["puncta_line_mode"], "green_puncta")
        self.assertEqual(defaults["nuclear_cell_pair_mode"], "red_nucleus")
        self.assertEqual(defaults["nuclear_cell_pair_contour_mode"], "aggressive")
        self.assertEqual(defaults["microns_per_pixel"], 0.25)
        self.assertFalse(defaults["use_metadata_scale"])
        self.assertFalse(defaults["use_metadata_channel_order"])
        self.assertEqual(
            defaults["fallback_channel_order"],
            [CHANNEL_ROLE_GREEN, CHANNEL_ROLE_DIC, CHANNEL_ROLE_RED, CHANNEL_ROLE_BLUE],
        )
        self.assertEqual(defaults["spatial_stats_unit"], "px")

    def test_experiment_workflow_defaults_endpoint_preserves_non_popup_preferences(
        self,
    ):
        preferences = get_user_preferences(self.user)
        preferences["auto_save_experiments"] = False
        preferences["show_saved_file_channels"] = False
        preferences["show_saved_file_scales"] = False
        preferences["sidebar_starts_open"] = False
        preferences["confirm_cell_deletion"] = False
        preferences["confirm_multi_cell_deletion"] = False
        preferences["sidebar_spatial_stats_unit"] = "um"
        preferences["main_image_channel"] = "green"
        preferences["default_puncta_source_contour_count_filter"] = "exactly_2"
        preferences["experiment_defaults"]["spatial_stats_unit"] = "um"
        update_user_preferences(self.user, preferences)

        response = self.client.post(
            reverse("experiment_workflow_defaults"),
            data=json.dumps(self._build_experiment_workflow_defaults_payload()),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.user.refresh_from_db()
        updated = get_user_preferences(self.user)
        self.assertFalse(updated["auto_save_experiments"])
        self.assertFalse(updated["show_saved_file_channels"])
        self.assertFalse(updated["show_saved_file_scales"])
        self.assertFalse(updated["sidebar_starts_open"])
        self.assertFalse(updated["confirm_cell_deletion"])
        self.assertFalse(updated["confirm_multi_cell_deletion"])
        self.assertEqual(updated["sidebar_spatial_stats_unit"], "um")
        self.assertEqual(updated["main_image_channel"], "green")
        self.assertEqual(
            updated["default_puncta_source_contour_count_filter"],
            "exactly_2",
        )
        self.assertEqual(updated["experiment_defaults"]["spatial_stats_unit"], "um")

    def test_experiment_workflow_defaults_endpoint_rejects_invalid_payload(self):
        baseline = get_user_preferences(self.user)
        cases = (
            {"selected_plugins": ["UnknownPlugin"]},
            {"puncta_line_width_unit": "bad"},
            {"green_dot_split_mode": "invalid"},
            {"nuclear_cell_pair_contour_mode": "invalid"},
            {"manual_required_channels": ["DIC"]},
            {"fallback_channel_order": ["DIC", "DIC", "channel_red", "channel_green"]},
        )

        for overrides in cases:
            payload = self._build_experiment_workflow_defaults_payload()
            payload.update(overrides)
            response = self.client.post(
                reverse("experiment_workflow_defaults"),
                data=json.dumps(payload),
                content_type="application/json",
            )
            self.assertEqual(response.status_code, 400)
            self.assertTrue(response.json()["errors"])
            self.user.refresh_from_db()
            self.assertEqual(get_user_preferences(self.user), baseline)

    def test_experiment_page_uses_updated_saved_workflow_defaults(self):
        response = self.client.post(
            reverse("experiment_workflow_defaults"),
            data=json.dumps(self._build_experiment_workflow_defaults_payload()),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)

        experiment_response = self.client.get(reverse("experiment"))
        self.assertEqual(experiment_response.status_code, 200)
        rendered_defaults = json.loads(
            experiment_response.context["user_preference_defaults_json"]
        )
        self.assertEqual(
            rendered_defaults["selected_plugins"],
            ["PunctaDistance", "GreenRedIntensity"],
        )
        self.assertEqual(rendered_defaults["cen_dot_proximity_radius"], 6.5)
        self.assertEqual(rendered_defaults["green_dot_split_mode"], "aggressive")
        self.assertEqual(rendered_defaults["nuclear_cell_pair_mode"], "red_nucleus")
        self.assertEqual(
            rendered_defaults["nuclear_cell_pair_contour_mode"], "aggressive"
        )
        self.assertFalse(rendered_defaults["use_metadata_channel_order"])
        self.assertEqual(
            rendered_defaults["fallback_channel_order"],
            [CHANNEL_ROLE_GREEN, CHANNEL_ROLE_DIC, CHANNEL_ROLE_RED, CHANNEL_ROLE_BLUE],
        )

    def test_experiment_page_restores_saved_nuclear_signal_quantification_defaults(
        self,
    ):
        payload = self._build_experiment_workflow_defaults_payload()
        payload.update(
            {
                "selected_plugins": ["CENDot"],
                "signal_quantification_enabled": True,
                "signal_quantification_mode": "nuclear_cell_pair",
                "puncta_contour_intensity_enabled": False,
                "alternate_nucleus_detection_enabled": True,
                "alternate_red_detection": True,
                "nuclear_cell_pair_mode": "green_nucleus",
            }
        )

        response = self.client.post(
            reverse("experiment_workflow_defaults"),
            data=json.dumps(payload),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)

        experiment_response = self.client.get(reverse("experiment"))
        self.assertEqual(experiment_response.status_code, 200)
        rendered_defaults = json.loads(
            experiment_response.context["user_preference_defaults_json"]
        )
        self.assertEqual(
            rendered_defaults["selected_plugins"],
            ["CENDot", "NuclearCellPairIntensity"],
        )
        self.assertTrue(rendered_defaults["signal_quantification_enabled"])
        self.assertEqual(
            rendered_defaults["signal_quantification_mode"], "nuclear_cell_pair"
        )
        self.assertFalse(rendered_defaults["puncta_contour_intensity_enabled"])
        self.assertTrue(rendered_defaults["alternate_nucleus_detection_enabled"])
        self.assertEqual(rendered_defaults["nuclear_cell_pair_mode"], "green_nucleus")

    def test_preferences_plugin_payload_includes_exclusive_and_dependency_fields(self):
        response = self.client.get(reverse("workflow_defaults"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '"required_plugins"', html=False)
        self.assertContains(response, '"exclusive_group"', html=False)
        self.assertContains(
            response, '"exclusive_group": "nuclear_cell_pair"', html=False
        )

    def test_advanced_settings_override_reports_and_removes_dependent_plugins(self):
        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_advanced_settings",
                "override_required_channels": ["channel_red"],
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response, "Advanced settings saved. Removed dependent plugins:"
        )
        for plugin_id in (
            "PunctaDistance",
            "CENDot",
            "Biorientation",
            "GreenRedIntensity",
        ):
            self.assertContains(response, PLUGIN_DEFINITIONS[plugin_id].label)

        self.user.refresh_from_db()
        defaults = get_user_preferences(self.user)["experiment_defaults"]
        self.assertEqual(defaults["selected_plugins"], [])

    def test_dashboard_channel_visibility_requires_boolean(self):
        response = self.client.post(
            reverse("dashboard_channel_visibility"),
            data=json.dumps({"show_saved_file_channels": "yes"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 400)

    def test_dashboard_scale_visibility_requires_boolean(self):
        response = self.client.post(
            reverse("dashboard_channel_visibility"),
            data=json.dumps({"show_saved_file_scales": "yes"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 400)

    def test_dashboard_sidebar_spatial_unit_requires_valid_unit(self):
        response = self.client.post(
            reverse("dashboard_channel_visibility"),
            data=json.dumps({"sidebar_spatial_stats_unit": "bad"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 400)

    def test_dashboard_main_image_channel_requires_valid_channel(self):
        response = self.client.post(
            reverse("dashboard_channel_visibility"),
            data=json.dumps({"main_image_channel": "purple"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 400)

    def test_dashboard_channel_visibility_persists_user_preference(self):
        response = self.client.post(
            reverse("dashboard_channel_visibility"),
            data=json.dumps({"show_saved_file_channels": False}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.user.refresh_from_db()
        self.assertFalse(get_user_preferences(self.user)["show_saved_file_channels"])

    def test_dashboard_scale_visibility_persists_user_preference(self):
        response = self.client.post(
            reverse("dashboard_channel_visibility"),
            data=json.dumps({"show_saved_file_scales": False}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.user.refresh_from_db()
        self.assertFalse(get_user_preferences(self.user)["show_saved_file_scales"])

    def test_dashboard_sidebar_spatial_unit_persists_without_changing_workflow_default(
        self,
    ):
        prefs = get_user_preferences(self.user)
        prefs["experiment_defaults"]["spatial_stats_unit"] = "px"
        update_user_preferences(self.user, prefs)

        response = self.client.post(
            reverse("dashboard_channel_visibility"),
            data=json.dumps({"sidebar_spatial_stats_unit": "um"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["sidebar_spatial_stats_unit"], "um")

        self.user.refresh_from_db()
        updated = get_user_preferences(self.user)
        self.assertEqual(updated["sidebar_spatial_stats_unit"], "um")
        self.assertEqual(updated["experiment_defaults"]["spatial_stats_unit"], "px")

    def test_dashboard_main_image_channel_persists_user_preference(self):
        response = self.client.post(
            reverse("dashboard_channel_visibility"),
            data=json.dumps({"main_image_channel": "green"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["main_image_channel"], "green")
        self.user.refresh_from_db()
        self.assertEqual(get_user_preferences(self.user)["main_image_channel"], "green")

    def test_behavior_form_persists_channel_visibility_toggle(self):
        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_behavior",
                "auto_save_experiments": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"], f"{reverse('workflow_defaults')}?section=saving"
        )
        self.user.refresh_from_db()
        self.assertFalse(get_user_preferences(self.user)["show_saved_file_channels"])
        self.assertFalse(get_user_preferences(self.user)["show_saved_file_scales"])

    def test_behavior_form_persists_sidebar_start_preference(self):
        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_behavior",
                "auto_save_experiments": "on",
                "show_saved_file_channels": "on",
                "show_saved_file_scales": "on",
                "sidebar_starts_open": "on",
                "default_puncta_source_contour_count_filter": "all",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(get_user_preferences(self.user)["sidebar_starts_open"])

    def test_behavior_form_persists_default_source_contour_count_filter(self):
        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_behavior",
                "auto_save_experiments": "on",
                "show_saved_file_channels": "on",
                "show_saved_file_scales": "on",
                "default_puncta_source_contour_count_filter": "exactly_1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.user.refresh_from_db()
        self.assertEqual(
            get_user_preferences(self.user)[
                "default_puncta_source_contour_count_filter"
            ],
            "exactly_1",
        )

        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_behavior",
                "auto_save_experiments": "on",
                "show_saved_file_channels": "on",
                "show_saved_file_scales": "on",
                "default_puncta_source_contour_count_filter": "",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.user.refresh_from_db()
        self.assertEqual(
            get_user_preferences(self.user)[
                "default_puncta_source_contour_count_filter"
            ],
            "all",
        )

    def test_behavior_form_preserves_default_source_filter_when_field_missing(self):
        preferences = get_user_preferences(self.user)
        preferences["default_puncta_source_contour_count_filter"] = "exactly_2"
        update_user_preferences(self.user, preferences)

        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_behavior",
                "auto_save_experiments": "on",
                "show_saved_file_channels": "on",
                "show_saved_file_scales": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.user.refresh_from_db()
        self.assertEqual(
            get_user_preferences(self.user)[
                "default_puncta_source_contour_count_filter"
            ],
            "exactly_2",
        )

    def test_behavior_form_persists_cell_delete_confirmation_preference(self):
        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_behavior",
                "auto_save_experiments": "on",
                "show_saved_file_channels": "on",
                "show_saved_file_scales": "on",
                "sidebar_starts_open": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.user.refresh_from_db()
        self.assertFalse(get_user_preferences(self.user)["confirm_cell_deletion"])

        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_behavior",
                "auto_save_experiments": "on",
                "show_saved_file_channels": "on",
                "show_saved_file_scales": "on",
                "sidebar_starts_open": "on",
                "confirm_cell_deletion": "on",
                "default_puncta_source_contour_count_filter": "all",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.user.refresh_from_db()
        self.assertTrue(get_user_preferences(self.user)["confirm_cell_deletion"])

    def test_behavior_form_persists_multi_cell_delete_confirmation_preference(self):
        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_behavior",
                "auto_save_experiments": "on",
                "show_saved_file_channels": "on",
                "show_saved_file_scales": "on",
                "sidebar_starts_open": "on",
                "confirm_cell_deletion": "on",
                "default_puncta_source_contour_count_filter": "all",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.user.refresh_from_db()
        self.assertFalse(get_user_preferences(self.user)["confirm_multi_cell_deletion"])

        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_behavior",
                "auto_save_experiments": "on",
                "show_saved_file_channels": "on",
                "show_saved_file_scales": "on",
                "sidebar_starts_open": "on",
                "confirm_cell_deletion": "on",
                "confirm_multi_cell_deletion": "on",
                "default_puncta_source_contour_count_filter": "all",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.user.refresh_from_db()
        self.assertTrue(get_user_preferences(self.user)["confirm_multi_cell_deletion"])

    def test_preferences_page_renders_cell_delete_confirmation_toggle(self):
        response = self.client.get(reverse("workflow_defaults") + "?section=saving")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="confirm_cell_deletion"', html=False)
        self.assertContains(response, 'id="confirm_multi_cell_deletion"', html=False)
        self.assertContains(
            response, 'data-workflow-card="deletion-preferences"', html=False
        )
        self.assertContains(response, "Deletion Preferences")
        self.assertContains(response, "Confirm Before Deleting Cells")
        self.assertContains(response, "Confirm Before Deleting Multiple Cells")

    def test_preferences_page_renders_result_display_default_filter(self):
        response = self.client.get(reverse("workflow_defaults") + "?section=saving")

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            'data-workflow-card="result-display-defaults"',
            html=False,
        )
        self.assertContains(response, "Result Display Defaults")
        self.assertContains(response, "Default Source Contour Count Filter")
        self.assertContains(response, "Include cells by default:")
        self.assertContains(
            response,
            'id="default_puncta_source_contour_count_filter"',
            html=False,
        )
        self.assertContains(response, '<option value="all"', html=False)
        self.assertContains(response, '<option value="exactly_1"', html=False)
        self.assertContains(response, '<option value="exactly_2"', html=False)
        self.assertContains(
            response,
            "Sets the default filter used when opening Display and Dashboard results. "
            "You can still change this filter on each results page.",
        )

    def test_behavior_form_honors_safe_next_redirect(self):
        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_behavior",
                "next": "/dashboard/",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/dashboard/")

    def test_behavior_form_rejects_external_next_redirect(self):
        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_behavior",
                "next": "https://example.com/phish",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"], f"{reverse('workflow_defaults')}?section=saving"
        )

    def test_behavior_form_disables_auto_save_when_toggle_is_off(self):
        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_behavior",
                "show_saved_file_channels": "on",
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            "Experiment autosave disabled. New runs will stay out of your dashboard history.",
        )

        self.user.refresh_from_db()
        preferences = get_user_preferences(self.user)
        self.assertFalse(preferences["auto_save_experiments"])
        self.assertTrue(preferences["show_saved_file_channels"])
        self.assertFalse(preferences["show_saved_file_scales"])
        self.assertFalse(should_auto_save_experiments(self.user))

    def test_behavior_form_enables_auto_save_when_toggle_is_on(self):
        existing = get_user_preferences(self.user)
        existing["auto_save_experiments"] = False
        update_user_preferences(self.user, existing)
        self.assertFalse(should_auto_save_experiments(self.user))

        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_behavior",
                "auto_save_experiments": "on",
                "show_saved_file_channels": "on",
                "show_saved_file_scales": "on",
                "sidebar_starts_open": "on",
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            "Experiment autosave enabled. New runs will appear on your dashboard.",
        )

        self.user.refresh_from_db()
        preferences = get_user_preferences(self.user)
        self.assertTrue(preferences["auto_save_experiments"])
        self.assertTrue(preferences["show_saved_file_channels"])
        self.assertTrue(preferences["show_saved_file_scales"])
        self.assertTrue(preferences["sidebar_starts_open"])
        self.assertTrue(should_auto_save_experiments(self.user))

    def test_dashboard_display_and_preprocess_sidebars_start_open_by_default(self):
        saved_uuid = self._create_saved_sidebar_file()
        preprocess_uuid = self._create_preprocess_sidebar_file()

        dashboard_response = self.client.get(reverse("dashboard"))
        display_response = self.client.get(reverse("display", args=[saved_uuid]))
        preprocess_response = self.client.get(
            reverse("pre_process", args=[preprocess_uuid])
        )

        self.assertEqual(dashboard_response.status_code, 200)
        self.assertEqual(display_response.status_code, 200)
        self.assertEqual(preprocess_response.status_code, 200)
        self.assertNotContains(
            dashboard_response, 'class="sidebar collapsed"', html=False
        )
        self.assertNotContains(
            display_response, 'class="sidebar collapsed"', html=False
        )
        self.assertNotContains(
            preprocess_response, 'class="sidebar collapsed"', html=False
        )

    def test_dashboard_display_and_preprocess_sidebars_render_collapsed_when_preference_is_off(
        self,
    ):
        prefs = get_user_preferences(self.user)
        prefs["sidebar_starts_open"] = False
        update_user_preferences(self.user, prefs)

        saved_uuid = self._create_saved_sidebar_file(filename="sidebar_saved_closed")
        preprocess_uuid = self._create_preprocess_sidebar_file(
            filename="sidebar_preprocess_closed"
        )

        dashboard_response = self.client.get(reverse("dashboard"))
        display_response = self.client.get(reverse("display", args=[saved_uuid]))
        preprocess_response = self.client.get(
            reverse("pre_process", args=[preprocess_uuid])
        )

        self.assertEqual(dashboard_response.status_code, 200)
        self.assertEqual(display_response.status_code, 200)
        self.assertEqual(preprocess_response.status_code, 200)
        self.assertContains(dashboard_response, 'class="sidebar collapsed"', html=False)
        self.assertContains(display_response, 'class="sidebar collapsed"', html=False)
        self.assertContains(
            preprocess_response, 'class="sidebar collapsed"', html=False
        )

    def test_new_user_has_default_selected_plugins(self):
        preferences = get_user_preferences(self.user)
        defaults = preferences["experiment_defaults"]
        self.assertEqual(
            defaults["selected_plugins"],
            [
                "PunctaDistance",
                "CENDot",
                "Biorientation",
                "GreenRedIntensity",
            ],
        )
        self.assertTrue(defaults["signal_quantification_enabled"])
        self.assertEqual(defaults["signal_quantification_mode"], "puncta_distance")
        self.assertTrue(defaults["puncta_contour_intensity_enabled"])
        self.assertEqual(defaults["puncta_line_mode"], "red_puncta")
        self.assertEqual(defaults["nuclear_cell_pair_mode"], "green_nucleus")
        self.assertTrue(defaults["green_dot_split_enabled"])
        self.assertEqual(defaults["green_dot_split_mode"], "balanced")
        self.assertNotIn("puncta_source_contour_count_filter", defaults)
        self.assertEqual(
            preferences["default_puncta_source_contour_count_filter"],
            "all",
        )

    def test_plugin_settings_form_persists_measurement_defaults(self):
        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_plugin_defaults",
                "selected_plugins": ["PunctaDistance"],
                "puncta_line_width": "2.5",
                "puncta_line_width_unit": "um",
                "cen_dot_distance": "11.2",
                "cen_dot_distance_unit": "px",
                "biorientation_red_min_distance": "2.5",
                "biorientation_red_min_distance_unit": "px",
                "biorientation_red_max_distance": "44.5",
                "biorientation_red_max_distance_unit": "um",
                "biorientation_collinearity_threshold": "77",
                "puncta_line_mode": "green_puncta",
                "nuclear_cell_pair_mode": "red_nucleus",
                "green_contour_filter_enabled": "on",
                "green_dot_split_enabled": "0",
                "green_dot_split_mode": "aggressive",
                "alternate_red_detection": "on",
                "microns_per_pixel": "0.25",
                "use_metadata_scale": "on",
                "spatial_stats_unit": "um",
                "use_metadata_channel_order": "0",
                "fallback_channel_order": [
                    "channel_green",
                    "DIC",
                    "channel_red",
                    "channel_blue",
                ],
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"], f"{reverse('workflow_defaults')}?section=plugins"
        )

        self.user.refresh_from_db()
        defaults = get_user_preferences(self.user)["experiment_defaults"]
        self.assertEqual(defaults["selected_plugins"], ["PunctaDistance"])
        self.assertEqual(defaults["puncta_line_width"], 2.5)
        self.assertEqual(defaults["puncta_line_width_unit"], "um")
        self.assertEqual(defaults["cen_dot_distance"], 11.2)
        self.assertEqual(defaults["cen_dot_distance_unit"], "px")
        self.assertEqual(defaults["biorientation_red_min_distance"], 2.5)
        self.assertEqual(defaults["biorientation_red_min_distance_unit"], "px")
        self.assertEqual(defaults["biorientation_red_max_distance"], 44.5)
        self.assertEqual(defaults["biorientation_red_max_distance_unit"], "um")
        self.assertEqual(defaults["biorientation_collinearity_threshold"], 77)
        self.assertTrue(defaults["green_contour_filter_enabled"])
        self.assertFalse(defaults["green_dot_split_enabled"])
        self.assertEqual(defaults["green_dot_split_mode"], "aggressive")
        self.assertNotIn("puncta_source_contour_count_filter", defaults)
        self.assertTrue(defaults["alternate_nucleus_detection_enabled"])
        self.assertTrue(defaults["alternate_red_detection"])
        self.assertEqual(defaults["puncta_line_mode"], "green_puncta")
        self.assertEqual(defaults["nuclear_cell_pair_mode"], "red_nucleus")
        self.assertEqual(defaults["microns_per_pixel"], 0.25)
        self.assertTrue(defaults["use_metadata_scale"])
        self.assertEqual(defaults["spatial_stats_unit"], "um")
        self.assertFalse(defaults["use_metadata_channel_order"])
        self.assertEqual(
            defaults["fallback_channel_order"],
            [CHANNEL_ROLE_GREEN, CHANNEL_ROLE_DIC, CHANNEL_ROLE_RED, CHANNEL_ROLE_BLUE],
        )

    def test_plugin_settings_form_preserves_paused_secondary_plugins_in_nuclear_mode(
        self,
    ):
        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_plugin_defaults",
                "selected_plugins": [
                    "NuclearCellPairIntensity",
                    "CENDot",
                    "Biorientation",
                ],
                "signal_quantification_enabled": "1",
                "signal_quantification_mode": "nuclear_cell_pair",
                "puncta_contour_intensity_enabled": "1",
            },
        )
        self.assertEqual(response.status_code, 302)

        self.user.refresh_from_db()
        defaults = get_user_preferences(self.user)["experiment_defaults"]
        self.assertEqual(
            defaults["selected_plugins"],
            ["CENDot", "Biorientation", "NuclearCellPairIntensity"],
        )
        selection = resolve_signal_quantification_selection(
            payload=defaults,
            selected_plugins=defaults["selected_plugins"],
            nuclear_cell_pair_mode=defaults["nuclear_cell_pair_mode"],
        )
        self.assertEqual(selection.selected_plugins, ("NuclearCellPairIntensity",))

    def test_advanced_settings_save_preserves_measurement_defaults(self):
        payload = get_user_preferences(self.user)
        payload["experiment_defaults"].update(
            {
                "puncta_line_width": 3.5,
                "puncta_line_width_unit": "um",
                "cen_dot_distance": 9.0,
                "cen_dot_distance_unit": "um",
                "biorientation_red_min_distance": 2,
                "biorientation_red_min_distance_unit": "px",
                "biorientation_red_max_distance": 41,
                "biorientation_red_max_distance_unit": "um",
                "biorientation_collinearity_threshold": 81,
                "green_contour_filter_enabled": True,
                "green_dot_split_enabled": False,
                "green_dot_split_mode": "aggressive",
                "alternate_nucleus_detection_enabled": True,
                "alternate_red_detection": True,
                "puncta_line_mode": "green_puncta",
                "nuclear_cell_pair_mode": "red_nucleus",
                "microns_per_pixel": 0.33,
                "use_metadata_scale": False,
                "spatial_stats_unit": "um",
                "use_metadata_channel_order": False,
                "fallback_channel_order": [
                    CHANNEL_ROLE_GREEN,
                    CHANNEL_ROLE_DIC,
                    CHANNEL_ROLE_RED,
                    CHANNEL_ROLE_BLUE,
                ],
            }
        )
        update_user_preferences(self.user, payload)

        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_advanced_settings",
                "module_enabled": "on",
                "enforce_layer_count": "on",
                "enforce_wavelengths": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"], f"{reverse('workflow_defaults')}?section=advanced"
        )

        self.user.refresh_from_db()
        defaults = get_user_preferences(self.user)["experiment_defaults"]
        self.assertEqual(defaults["puncta_line_width"], 3.5)
        self.assertEqual(defaults["puncta_line_width_unit"], "um")
        self.assertEqual(defaults["cen_dot_distance"], 9.0)
        self.assertEqual(defaults["cen_dot_distance_unit"], "um")
        self.assertEqual(defaults["biorientation_red_min_distance"], 2)
        self.assertEqual(defaults["biorientation_red_min_distance_unit"], "px")
        self.assertEqual(defaults["biorientation_red_max_distance"], 41)
        self.assertEqual(defaults["biorientation_red_max_distance_unit"], "um")
        self.assertEqual(defaults["biorientation_collinearity_threshold"], 81)
        self.assertTrue(defaults["green_contour_filter_enabled"])
        self.assertFalse(defaults["green_dot_split_enabled"])
        self.assertEqual(defaults["green_dot_split_mode"], "aggressive")
        self.assertTrue(defaults["alternate_red_detection"])
        self.assertEqual(defaults["puncta_line_mode"], "green_puncta")
        self.assertEqual(defaults["nuclear_cell_pair_mode"], "red_nucleus")
        self.assertEqual(defaults["microns_per_pixel"], 0.33)
        self.assertFalse(defaults["use_metadata_scale"])
        self.assertEqual(defaults["spatial_stats_unit"], "um")
        self.assertFalse(defaults["use_metadata_channel_order"])
        self.assertEqual(
            defaults["fallback_channel_order"],
            [CHANNEL_ROLE_GREEN, CHANNEL_ROLE_DIC, CHANNEL_ROLE_RED, CHANNEL_ROLE_BLUE],
        )

    def test_plugin_settings_form_persists_dot_split_defaults(self):
        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_plugin_defaults",
                "selected_plugins": ["PunctaDistance"],
                "green_dot_split_enabled": "0",
                "green_dot_split_mode": "aggressive",
                "red_dot_split_enabled": "1",
                "red_dot_split_mode": "aggressive",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"], f"{reverse('workflow_defaults')}?section=plugins"
        )

        self.user.refresh_from_db()
        defaults = get_user_preferences(self.user)["experiment_defaults"]
        self.assertFalse(defaults["green_dot_split_enabled"])
        self.assertEqual(defaults["green_dot_split_mode"], "aggressive")
        self.assertTrue(defaults["red_dot_split_enabled"])
        self.assertEqual(defaults["red_dot_split_mode"], "aggressive")

        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_plugin_defaults",
                "selected_plugins": ["PunctaDistance"],
                "green_dot_split_enabled": "on",
                "green_dot_split_mode": "invalid",
                "red_dot_split_enabled": "0",
                "red_dot_split_mode": "invalid",
            },
        )
        self.assertEqual(response.status_code, 302)

        self.user.refresh_from_db()
        defaults = get_user_preferences(self.user)["experiment_defaults"]
        self.assertTrue(defaults["green_dot_split_enabled"])
        self.assertEqual(defaults["green_dot_split_mode"], "balanced")
        self.assertFalse(defaults["red_dot_split_enabled"])
        self.assertEqual(defaults["red_dot_split_mode"], "balanced")

    def test_advanced_settings_pauses_optional_checks_when_module_disabled(self):
        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_advanced_settings",
                "enforce_layer_count": "on",
                "enforce_wavelengths": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"], f"{reverse('workflow_defaults')}?section=advanced"
        )

        self.user.refresh_from_db()
        defaults = get_user_preferences(self.user)["experiment_defaults"]
        self.assertFalse(defaults["module_enabled"])
        self.assertTrue(defaults["enforce_layer_count"])
        self.assertTrue(defaults["enforce_wavelengths"])

    def test_workflow_defaults_summary_marks_manual_channels_active_in_both_sections(
        self,
    ):
        preferences = get_user_preferences(self.user)
        preferences["experiment_defaults"]["module_enabled"] = True
        preferences["experiment_defaults"]["manual_required_channels"] = [
            "channel_blue"
        ]
        update_user_preferences(self.user, preferences)

        response = self.client.get(reverse("workflow_defaults"))
        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertEqual(content.count(">Required manually<"), 2)
        self.assertIn(
            'data-req-row="channel_blue" data-summary-scope="plugins"', content
        )
        self.assertIn(
            'data-req-row="channel_blue" data-summary-scope="advanced"', content
        )

    def test_workflow_defaults_summary_marks_manual_channels_paused_in_both_sections(
        self,
    ):
        preferences = get_user_preferences(self.user)
        preferences["experiment_defaults"]["module_enabled"] = False
        preferences["experiment_defaults"]["manual_required_channels"] = [
            "channel_blue"
        ]
        update_user_preferences(self.user, preferences)

        response = self.client.get(reverse("workflow_defaults"))
        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertEqual(content.count(">Paused manually<"), 2)
        self.assertContains(response, 'id="manual_channel_blue"', html=False)
        self.assertContains(response, 'value="channel_blue"', html=False)
        self.assertContains(response, 'data-channel="channel_blue"', html=False)
        self.assertContains(response, "checked disabled", html=False)

    def test_workflow_defaults_summary_marks_paused_all_wavelengths_in_both_sections(
        self,
    ):
        preferences = get_user_preferences(self.user)
        preferences["experiment_defaults"]["module_enabled"] = False
        preferences["experiment_defaults"]["enforce_wavelengths"] = True
        update_user_preferences(self.user, preferences)

        response = self.client.get(reverse("workflow_defaults"))
        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertEqual(content.count(">Paused by all-channels<"), 2)

    def test_workflow_defaults_renders_optional_validation_controls_disabled_when_module_is_off(
        self,
    ):
        preferences = get_user_preferences(self.user)
        preferences["experiment_defaults"]["module_enabled"] = False
        preferences["experiment_defaults"]["enforce_layer_count"] = True
        preferences["experiment_defaults"]["enforce_wavelengths"] = True
        preferences["experiment_defaults"]["manual_required_channels"] = [
            "channel_blue"
        ]
        update_user_preferences(self.user, preferences)

        response = self.client.get(reverse("workflow_defaults"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="advancedOptionalChecksGroup"', html=False)
        self.assertContains(
            response, 'id="enforce_layer_count" checked disabled', html=False
        )
        self.assertContains(
            response, 'id="enforce_wavelengths" checked disabled', html=False
        )
        self.assertContains(response, 'id="manual_channel_blue"', html=False)
        self.assertContains(response, 'value="channel_blue"', html=False)
        self.assertContains(response, 'data-channel="channel_blue"', html=False)
        self.assertContains(response, "checked disabled", html=False)

    def test_advanced_settings_save_preserves_manual_required_channels_when_module_is_off(
        self,
    ):
        response = self.client.post(
            reverse("workflow_defaults"),
            {
                "action": "save_advanced_settings",
                "enforce_layer_count": "1",
                "manual_required_channels": ["channel_blue"],
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"], f"{reverse('workflow_defaults')}?section=advanced"
        )

        self.user.refresh_from_db()
        defaults = get_user_preferences(self.user)["experiment_defaults"]
        self.assertFalse(defaults["module_enabled"])
        self.assertTrue(defaults["enforce_layer_count"])
        self.assertFalse(defaults["enforce_wavelengths"])
        self.assertEqual(defaults["manual_required_channels"], ["channel_blue"])

        response = self.client.get(reverse("workflow_defaults"))
        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertEqual(content.count(">Paused manually<"), 2)
