"""Protect cell deletion routes, artifacts, and viewer payload cleanup."""

from __future__ import annotations

import json
import re
import csv
from io import BytesIO
from contextlib import ExitStack, contextmanager
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch
from uuid import uuid4

from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook
from PIL import Image

from core.config import DEFAULT_CHANNEL_CONFIG
from core.models import CellStatistics, SegmentedImage, UploadedImage
from core.services.cell_deletion import delete_multiple_cells, delete_single_cell


@contextmanager
def temporary_media_root():
    with TemporaryDirectory() as temp_media:
        with ExitStack() as stack:
            stack.enter_context(override_settings(MEDIA_ROOT=temp_media))
            for target in (
                "accounts.views.profile.MEDIA_ROOT",
                "core.config.MEDIA_ROOT",
                "core.views.display.MEDIA_ROOT",
            ):
                stack.enter_context(patch(target, temp_media))
            yield Path(temp_media)


def _write_channel_config(media_root: Path, uuid_value: str) -> None:
    run_dir = media_root / uuid_value
    run_dir.mkdir(parents=True, exist_ok=True)
    (run_dir / "channel_config.json").write_text(
        json.dumps(DEFAULT_CHANNEL_CONFIG),
        encoding="utf-8",
    )


def _create_cell_artifacts(
    media_root: Path, uuid_value: str, dv_name: str, cell_id: int
) -> dict[str, Path]:
    """Create the per-cell on-disk artifacts that delete_single_cell should remove."""

    segmented_dir = media_root / uuid_value / "segmented"
    segmented_dir.mkdir(parents=True, exist_ok=True)
    overlay_dir = segmented_dir / "overlay-cache-v4"
    overlay_dir.mkdir(parents=True, exist_ok=True)

    def _png(path: Path) -> Path:
        Image.new("RGB", (4, 4), color=(0, 0, 0)).save(path, format="PNG")
        return path

    paths: dict[str, Path] = {
        "binary_mask": _png(segmented_dir / f"cell_{cell_id}.png"),
        "outline": segmented_dir / f"{dv_name}-{cell_id}.outline",
    }
    paths["outline"].write_text("0,0\n", encoding="utf-8")
    for channel_index in DEFAULT_CHANNEL_CONFIG.values():
        idx = int(channel_index)
        paths[f"channel_{idx}"] = _png(
            segmented_dir / f"{dv_name}-{idx}-{cell_id}.png"
        )
        paths[f"channel_{idx}_no_outline"] = _png(
            segmented_dir / f"{dv_name}-{idx}-{cell_id}-no_outline.png"
        )
    for channel in ("red", "green", "blue"):
        paths[f"overlay_{channel}"] = _png(
            overlay_dir / f"cell-{cell_id}-{channel}.png"
        )
    paths["overlay_lock"] = overlay_dir / f"cell-{cell_id}.lock"
    paths["overlay_lock"].write_text("", encoding="utf-8")
    return paths


def _make_cell_stats(
    segmented: SegmentedImage,
    cell_id: int,
    dv_name: str,
) -> CellStatistics:
    return CellStatistics.objects.create(
        segmented_image=segmented,
        cell_id=cell_id,
        puncta_distance=0.0,
        puncta_line_intensity=0.0,
        nucleus_intensity_sum=0.0,
        cell_pair_intensity_sum=0.0,
        red_in_red_total_intensity_1=0.0,
        red_in_red_total_intensity_2=0.0,
        red_in_red_total_intensity_3=0.0,
        green_in_red_total_intensity_1=0.0,
        green_in_red_total_intensity_2=0.0,
        green_in_red_total_intensity_3=0.0,
        red_in_green_total_intensity_1=0.0,
        red_in_green_total_intensity_2=0.0,
        red_in_green_total_intensity_3=0.0,
        green_in_green_total_intensity_1=0.0,
        green_in_green_total_intensity_2=0.0,
        green_in_green_total_intensity_3=0.0,
        green_red_intensity_1=0.0,
        green_red_intensity_2=0.0,
        green_red_intensity_3=0.0,
        dv_file_path=f"{segmented.UUID}/{dv_name}.dv",
        image_name=f"{dv_name}.dv",
    )


def _rendered_table_cell_ids(response) -> list[int]:
    try:
        table = response.context["cell_table"]
        return [int(row.record.cell_id) for row in table.rows]
    except (AttributeError, KeyError, TypeError, ValueError):
        pass

    body = response.content.decode("utf-8", errors="ignore")
    match = re.search(
        r'<table[^>]*id="celltable"[^>]*>.*?<tbody>(.*?)</tbody>',
        body,
        flags=re.S,
    )
    if not match:
        return []
    return [
        int(value)
        for value in re.findall(
            r"<tr[^>]*>\s*<td[^>]*>\s*(\d+)\s*</td>",
            match.group(1),
            flags=re.S,
        )
    ]


def _csv_cell_ids(response) -> list[int]:
    body = response.content.decode("utf-8", errors="ignore")
    ids: list[int] = []
    for row in body.splitlines()[1:]:
        first_value = row.split(",", 1)[0].strip()
        if first_value.isdigit():
            ids.append(int(first_value))
    return ids


def _xlsx_cell_ids(response) -> list[int]:
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    ids: list[int] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        value = row[0]
        if value is None:
            continue
        ids.append(int(value))
    return ids


class CellDeletionServiceTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            email="cell-delete-tests@example.com",
            password="TestPass123!",
        )

    def _setup_run(self, media_root: Path, *, num_cells: int = 4):
        uuid_value = str(uuid4())
        dv_name = "sample_PRJ"
        _write_channel_config(media_root, uuid_value)
        UploadedImage.objects.create(
            user=self.user,
            uuid=uuid_value,
            name=dv_name,
            file_location=f"{uuid_value}/{dv_name}.dv",
        )
        segmented = SegmentedImage.objects.create(
            user=self.user,
            UUID=uuid_value,
            file_location=f"user_{uuid_value}/{dv_name}.png",
            ImagePath=f"{uuid_value}/output/{dv_name}_frame_0.png",
            CellPairPrefix=f"{uuid_value}/segmented/cell_",
            NumCells=num_cells,
        )
        artifacts: dict[int, dict[str, Path]] = {}
        for cell_id in range(1, num_cells + 1):
            _make_cell_stats(segmented, cell_id, dv_name)
            artifacts[cell_id] = _create_cell_artifacts(
                media_root, uuid_value, dv_name, cell_id
            )
        return segmented, artifacts

    def test_delete_single_cell_removes_row_and_decrements_num_cells(self):
        with temporary_media_root() as media_root:
            segmented, _ = self._setup_run(media_root, num_cells=4)
            self.assertEqual(segmented.NumCells, 4)

            delete_single_cell(segmented, cell_id=2)

            self.assertFalse(
                CellStatistics.objects
                .filter(segmented_image=segmented, cell_id=2)
                .exists()
            )
            segmented.refresh_from_db(fields=["NumCells"])
            self.assertEqual(segmented.NumCells, 3)
            remaining_ids = sorted(
                CellStatistics.objects
                .filter(segmented_image=segmented)
                .values_list("cell_id", flat=True)
            )
            self.assertEqual(remaining_ids, [1, 3, 4])

    def test_delete_single_cell_removes_on_disk_artifacts(self):
        with temporary_media_root() as media_root:
            segmented, artifacts = self._setup_run(media_root, num_cells=3)
            target_paths = artifacts[2]
            for path in target_paths.values():
                self.assertTrue(path.exists(), f"expected {path} before delete")

            delete_single_cell(segmented, cell_id=2)

            for label, path in target_paths.items():
                self.assertFalse(path.exists(), f"{label} ({path}) should be removed")

    def test_delete_single_cell_leaves_other_cells_untouched(self):
        with temporary_media_root() as media_root:
            segmented, artifacts = self._setup_run(media_root, num_cells=4)

            delete_single_cell(segmented, cell_id=2)

            for keep_id in (1, 3, 4):
                self.assertTrue(
                    CellStatistics.objects
                    .filter(segmented_image=segmented, cell_id=keep_id)
                    .exists()
                )
                for path in artifacts[keep_id].values():
                    self.assertTrue(
                        path.exists(),
                        f"cell {keep_id} artifact {path} should remain",
                    )

    def test_delete_single_cell_raises_for_missing_cell(self):
        with temporary_media_root() as media_root:
            segmented, _ = self._setup_run(media_root, num_cells=2)
            with self.assertRaises(CellStatistics.DoesNotExist):
                delete_single_cell(segmented, cell_id=999)

    def test_delete_multiple_cells_removes_rows_artifacts_and_updates_num_cells(self):
        with temporary_media_root() as media_root:
            segmented, artifacts = self._setup_run(media_root, num_cells=4)

            deleted_ids = delete_multiple_cells(segmented, [2, 3, 999])

            self.assertEqual(deleted_ids, [2, 3])
            segmented.refresh_from_db(fields=["NumCells"])
            self.assertEqual(segmented.NumCells, 2)
            remaining_ids = sorted(
                CellStatistics.objects
                .filter(segmented_image=segmented)
                .values_list("cell_id", flat=True)
            )
            self.assertEqual(remaining_ids, [1, 4])
            for deleted_id in (2, 3):
                for path in artifacts[deleted_id].values():
                    self.assertFalse(path.exists(), f"{path} should be removed")
            for keep_id in (1, 4):
                for path in artifacts[keep_id].values():
                    self.assertTrue(path.exists(), f"{path} should remain")

    def test_delete_multiple_cells_returns_empty_for_missing_cells(self):
        with temporary_media_root() as media_root:
            segmented, _ = self._setup_run(media_root, num_cells=2)

            deleted_ids = delete_multiple_cells(segmented, [99])

            self.assertEqual(deleted_ids, [])
            segmented.refresh_from_db(fields=["NumCells"])
            self.assertEqual(segmented.NumCells, 2)


class CellDeletionEndpointTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            email="endpoint-owner@example.com",
            password="TestPass123!",
        )
        self.other_user = user_model.objects.create_user(
            email="endpoint-other@example.com",
            password="TestPass123!",
        )

    def _setup_run(self, media_root: Path, *, owner=None, num_cells: int = 3):
        owner = owner or self.user
        uuid_value = str(uuid4())
        dv_name = "sample_PRJ"
        _write_channel_config(media_root, uuid_value)
        UploadedImage.objects.create(
            user=owner,
            uuid=uuid_value,
            name=dv_name,
            file_location=f"{uuid_value}/{dv_name}.dv",
        )
        segmented = SegmentedImage.objects.create(
            user=owner,
            UUID=uuid_value,
            file_location=f"user_{uuid_value}/{dv_name}.png",
            ImagePath=f"{uuid_value}/output/{dv_name}_frame_0.png",
            CellPairPrefix=f"{uuid_value}/segmented/cell_",
            NumCells=num_cells,
        )
        for cell_id in range(1, num_cells + 1):
            _make_cell_stats(segmented, cell_id, dv_name)
            _create_cell_artifacts(media_root, uuid_value, dv_name, cell_id)
        return uuid_value, segmented

    def test_endpoint_owner_can_delete_cell(self):
        with temporary_media_root() as media_root:
            uuid_value, segmented = self._setup_run(media_root, num_cells=3)
            self.client.login(
                email=self.user.email, password="TestPass123!"
            )
            url = reverse(
                "delete_cell", kwargs={"uuid": uuid_value, "cell_id": 2}
            )
            response = self.client.post(url)

            self.assertEqual(response.status_code, 200)
            payload = response.json()
            self.assertTrue(payload["ok"])
            self.assertEqual(payload["cell_id"], 2)
            self.assertEqual(payload["num_cells"], 2)
            self.assertEqual(payload["remaining_cells"], [1, 3])
            self.assertFalse(
                CellStatistics.objects
                .filter(segmented_image=segmented, cell_id=2)
                .exists()
            )

    def test_endpoint_rejects_non_owner(self):
        with temporary_media_root() as media_root:
            uuid_value, segmented = self._setup_run(media_root, num_cells=2)
            self.client.login(
                email=self.other_user.email, password="TestPass123!"
            )
            url = reverse(
                "delete_cell", kwargs={"uuid": uuid_value, "cell_id": 1}
            )
            response = self.client.post(url)

            self.assertEqual(response.status_code, 403)
            self.assertTrue(
                CellStatistics.objects
                .filter(segmented_image=segmented, cell_id=1)
                .exists()
            )

    def test_endpoint_returns_404_for_unknown_cell(self):
        with temporary_media_root() as media_root:
            uuid_value, _ = self._setup_run(media_root, num_cells=2)
            self.client.login(
                email=self.user.email, password="TestPass123!"
            )
            url = reverse(
                "delete_cell", kwargs={"uuid": uuid_value, "cell_id": 999}
            )
            response = self.client.post(url)
            self.assertEqual(response.status_code, 404)

    def test_bulk_endpoint_deletes_existing_cells_and_preserves_gaps(self):
        with temporary_media_root() as media_root:
            uuid_value, segmented = self._setup_run(media_root, num_cells=4)
            self.client.login(
                email=self.user.email, password="TestPass123!"
            )
            url = reverse("delete_cells", kwargs={"uuid": uuid_value})

            response = self.client.post(
                url,
                data=json.dumps({"cell_ids": [2, 3, 999]}),
                content_type="application/json",
            )

            self.assertEqual(response.status_code, 200)
            payload = response.json()
            self.assertTrue(payload["ok"])
            self.assertEqual(payload["deleted_cells"], [2, 3])
            self.assertEqual(payload["num_cells"], 2)
            self.assertEqual(payload["remaining_cells"], [1, 4])
            segmented.refresh_from_db(fields=["NumCells"])
            self.assertEqual(segmented.NumCells, 2)
            self.assertEqual(
                list(
                    CellStatistics.objects
                    .filter(segmented_image=segmented)
                    .order_by("cell_id")
                    .values_list("cell_id", flat=True)
                ),
                [1, 4],
            )

    def test_bulk_endpoint_rejects_invalid_payload(self):
        with temporary_media_root() as media_root:
            uuid_value, _ = self._setup_run(media_root, num_cells=2)
            self.client.login(
                email=self.user.email, password="TestPass123!"
            )
            url = reverse("delete_cells", kwargs={"uuid": uuid_value})

            response = self.client.post(
                url,
                data=json.dumps({"cell_ids": [1, "bad"]}),
                content_type="application/json",
            )

            self.assertEqual(response.status_code, 400)

    def test_bulk_endpoint_returns_404_when_no_requested_cells_exist(self):
        with temporary_media_root() as media_root:
            uuid_value, _ = self._setup_run(media_root, num_cells=2)
            self.client.login(
                email=self.user.email, password="TestPass123!"
            )
            url = reverse("delete_cells", kwargs={"uuid": uuid_value})

            response = self.client.post(
                url,
                data=json.dumps({"cell_ids": [99]}),
                content_type="application/json",
            )

            self.assertEqual(response.status_code, 404)

    def test_bulk_endpoint_rejects_non_owner(self):
        with temporary_media_root() as media_root:
            uuid_value, segmented = self._setup_run(media_root, num_cells=2)
            self.client.login(
                email=self.other_user.email, password="TestPass123!"
            )
            url = reverse("delete_cells", kwargs={"uuid": uuid_value})

            response = self.client.post(
                url,
                data=json.dumps({"cell_ids": [1, 2]}),
                content_type="application/json",
            )

            self.assertEqual(response.status_code, 403)
            self.assertEqual(
                CellStatistics.objects.filter(segmented_image=segmented).count(),
                2,
            )

    def test_csv_export_excludes_deleted_cell(self):
        with temporary_media_root() as media_root:
            uuid_value, _ = self._setup_run(media_root, num_cells=3)
            self.client.login(
                email=self.user.email, password="TestPass123!"
            )

            delete_url = reverse(
                "delete_cell", kwargs={"uuid": uuid_value, "cell_id": 2}
            )
            self.assertEqual(self.client.post(delete_url).status_code, 200)

            display_url = reverse("display", kwargs={"uuids": uuid_value})
            response = self.client.get(display_url + "?_export=csv&_unit=px")
            self.assertEqual(response.status_code, 200)
            body = response.content.decode("utf-8", errors="ignore")
            cell_id_column = [
                row.split(",", 1)[0].strip()
                for row in body.splitlines()[1:]
                if row.strip()
            ]
            self.assertNotIn("2", cell_id_column)
            self.assertIn("1", cell_id_column)
            self.assertIn("3", cell_id_column)

            xlsx_response = self.client.get(display_url + "?_export=xlsx&_unit=px")
            self.assertEqual(xlsx_response.status_code, 200)
            self.assertEqual(_xlsx_cell_ids(xlsx_response), [1, 3])

    def test_display_page_excludes_deleted_gap_cells_from_table_and_payload(self):
        with temporary_media_root() as media_root:
            uuid_value, _ = self._setup_run(media_root, num_cells=4)
            self.client.login(
                email=self.user.email, password="TestPass123!"
            )

            delete_url = reverse("delete_cells", kwargs={"uuid": uuid_value})
            response = self.client.post(
                delete_url,
                data=json.dumps({"cell_ids": [2, 3]}),
                content_type="application/json",
            )
            self.assertEqual(response.status_code, 200)

            response = self.client.get(reverse("display", kwargs={"uuids": uuid_value}))

            self.assertEqual(response.status_code, 200)
            self.assertEqual(_rendered_table_cell_ids(response), [1, 4])
            files_data = json.loads(response.context["files_data"])
            file_data = files_data[uuid_value]
            self.assertEqual(file_data["NumberOfCells"], 2)
            self.assertEqual(sorted(file_data["Statistics"].keys()), ["1", "4"])
            self.assertEqual(sorted(file_data["CellPairImages"].keys()), ["1", "4"])

    def test_dashboard_page_excludes_deleted_gap_cells_from_table_and_payload(self):
        with temporary_media_root() as media_root:
            uuid_value, _ = self._setup_run(media_root, num_cells=4)
            self.client.login(
                email=self.user.email, password="TestPass123!"
            )

            delete_url = reverse("delete_cells", kwargs={"uuid": uuid_value})
            response = self.client.post(
                delete_url,
                data=json.dumps({"cell_ids": [2, 3]}),
                content_type="application/json",
            )
            self.assertEqual(response.status_code, 200)

            response = self.client.get(reverse("dashboard"))

            self.assertEqual(response.status_code, 200)
            self.assertEqual(_rendered_table_cell_ids(response), [1, 4])
            files_data = json.loads(response.context["files_data_json"])
            file_data = files_data[uuid_value]
            self.assertEqual(file_data["NumberOfCells"], 2)
            self.assertEqual(sorted(file_data["Statistics"].keys()), ["1", "4"])
            self.assertEqual(sorted(file_data["CellPairImages"].keys()), ["1", "4"])
            body = response.content.decode("utf-8", errors="ignore")
            self.assertNotIn("for (let i = 1; i <= totalCells", body)

    def test_dashboard_exports_exclude_deleted_gap_cells(self):
        with temporary_media_root() as media_root:
            uuid_value, _ = self._setup_run(media_root, num_cells=4)
            self.client.login(
                email=self.user.email, password="TestPass123!"
            )

            delete_url = reverse("delete_cells", kwargs={"uuid": uuid_value})
            response = self.client.post(
                delete_url,
                data=json.dumps({"cell_ids": [2, 3]}),
                content_type="application/json",
            )
            self.assertEqual(response.status_code, 200)

            csv_response = self.client.get(
                reverse("dashboard"),
                {"file_uuid": uuid_value, "_export": "csv", "_unit": "px"},
            )
            xlsx_response = self.client.get(
                reverse("dashboard"),
                {"file_uuid": uuid_value, "_export": "xlsx", "_unit": "px"},
            )
            filtered_csv_response = self.client.get(
                reverse("dashboard"),
                {
                    "file_uuid": uuid_value,
                    "_export": "csv",
                    "_unit": "px",
                    "_columns": "puncta_distance",
                },
            )
            selected_intensity_csv_response = self.client.get(
                reverse("dashboard"),
                {
                    "file_uuid": uuid_value,
                    "_export": "csv",
                    "_unit": "px",
                    "_columns": "red_in_red_total_intensity_1",
                },
            )
            selected_intensity_xlsx_response = self.client.get(
                reverse("dashboard"),
                {
                    "file_uuid": uuid_value,
                    "_export": "xlsx",
                    "_unit": "px",
                    "_columns": "red_in_red_total_intensity_1",
                },
            )

            self.assertEqual(csv_response.status_code, 200)
            self.assertEqual(xlsx_response.status_code, 200)
            self.assertEqual(filtered_csv_response.status_code, 200)
            self.assertEqual(selected_intensity_csv_response.status_code, 200)
            self.assertEqual(selected_intensity_xlsx_response.status_code, 200)
            self.assertEqual(_csv_cell_ids(csv_response), [1, 4])
            self.assertEqual(_xlsx_cell_ids(xlsx_response), [1, 4])
            self.assertEqual(_csv_cell_ids(filtered_csv_response), [1, 4])
            self.assertEqual(_csv_cell_ids(selected_intensity_csv_response), [1, 4])
            self.assertEqual(_xlsx_cell_ids(selected_intensity_xlsx_response), [1, 4])

    def test_display_and_dashboard_exports_exclude_deleted_mixed_cell_type_rows(self):
        with temporary_media_root() as media_root:
            uuid_value, segmented = self._setup_run(media_root, num_cells=3)
            for cell_id, cell_type in ((1, "single_cell"), (2, "cell_pair"), (3, "cell_pair")):
                stat = CellStatistics.objects.get(
                    segmented_image=segmented,
                    cell_id=cell_id,
                )
                properties = dict(stat.properties or {})
                properties.update(
                    {
                        "cell_type": cell_type,
                        "signal_quantification_mode": "puncta_distance",
                        "puncta_line_mode": "red_puncta",
                    }
                )
                stat.cell_type = cell_type
                stat.properties = properties
                stat.save(update_fields=["cell_type", "properties"])
            self.client.login(
                email=self.user.email, password="TestPass123!"
            )

            delete_url = reverse("delete_cell", kwargs={"uuid": uuid_value, "cell_id": 2})
            self.assertEqual(self.client.post(delete_url).status_code, 200)

            urls = (
                reverse("display", kwargs={"uuids": uuid_value})
                + "?_export=csv&_unit=px&_columns=red_in_red_total_intensity_1",
                reverse("dashboard")
                + f"?file_uuid={uuid_value}&_export=csv&_unit=px"
                + "&_columns=red_in_red_total_intensity_1",
            )
            for url in urls:
                with self.subTest(url=url):
                    response = self.client.get(url)
                    self.assertEqual(response.status_code, 200)
                    rows = list(csv.reader(response.content.decode("utf-8").splitlines()))

                    self.assertEqual(
                        rows[0],
                        ["Cell ID", "Cell Type", "Red In Red Total Intensity 1"],
                    )
                    self.assertEqual(
                        [(row[0], row[1]) for row in rows[1:]],
                        [("1", "Single Cell"), ("3", "Cell Pair")],
                    )
                    self.assertNotIn(("2", "Cell Pair"), [(row[0], row[1]) for row in rows[1:]])
